#!/usr/bin/env python3
"""Create terminology-only Table S2 V5 and Table S3 V8 workbooks.

The source V4/V7 workbooks remain unchanged. Table S2 receives the corrected
BUSCO percentage and the first expansion of PEV. Table S3 uses affirmative
cross-representation language throughout its interpretation cells. All
pre-existing numeric cells must remain exactly identical.

Created: 2026-07-19 21:40:45 Asia/Bangkok
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment


SCRIPT_VERSION = "2026-07-19.1"
STAMP = "20260719_214300"
SCRIPT = Path(__file__).resolve()
SUPPLEMENT = SCRIPT.parent
ROOT = SCRIPT.parents[2]

S2_SOURCE = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_213408_V4.xlsx"
S2_SOURCE_INTEGRITY = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_213408_V4_integrity.json"
S3_SOURCE = SUPPLEMENT / "Table_S3_AEF_20260719_213408_V7.xlsx"
S3_SOURCE_INTEGRITY = SUPPLEMENT / "Table_S3_AEF_20260719_213408_V7_integrity.json"

S2_OUTPUT = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V5.xlsx"
S2_INTEGRITY = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V5_integrity.json"
S3_OUTPUT = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V8.xlsx"
S3_INTEGRITY = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V8_integrity.json"

EXPECTED_HASHES = {
    S2_SOURCE: "aed3fa23cb495484ec0da845a36f0069b2850880a26f5783b42c666586d4f219",
    S2_SOURCE_INTEGRITY: "a38be53165681305b52ffeaa716f6a55d3a64eef4c4cbe4266c78fbdb2c706e9",
    S3_SOURCE: "67d52948a4254d79b2f2a5cca48cc336d3f8dc53f19e258abf718b9b117015e3",
    S3_SOURCE_INTEGRITY: "738950fc701b55b57111ece802e9c75133591de3a740c8af34367f642c9c7333",
}

S2_REQUIRED_CHECKS = (
    "Unique-site raw counts; unique-site total-hit normalization; unique-site peptide "
    "normalization; quality/phylum site-cluster model; Benchmarking Universal Single-Copy "
    "Orthologs (BUSCO) ≥50% unique-site analysis; topology-aware site-cluster model using "
    "three phylogenetic eigenvector (PEV) covariates; and the structured null."
)

S3_AFFIRMATIVE_ALIGNMENT = (
    "descriptive site-level cross-representation alignment/crosswalk; axis meanings remain "
    "distributed and nonunique; both representations use the same recorded coordinates"
)

S3_AFFIRMATIVE_CANDIDATE = (
    "descriptive latent-feature association; biological interpretation remains candidate-level"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT))


def authenticate_inputs() -> None:
    for path, expected in EXPECTED_HASHES.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        observed = sha256(path)
        if observed != expected:
            raise RuntimeError(f"Authenticated source changed: {path}: {observed}")
    for integrity in (S2_SOURCE_INTEGRITY, S3_SOURCE_INTEGRITY):
        record = json.loads(integrity.read_text(encoding="utf-8"))
        if record.get("status") != "PASS" or record.get("audit", {}).get("result") != "PASS":
            raise RuntimeError(f"Source integrity record is not PASS: {integrity}")
    for output in (S2_OUTPUT, S2_INTEGRITY, S3_OUTPUT, S3_INTEGRITY):
        if output.exists():
            raise FileExistsError(f"Refusing to overwrite {output}")


def build_s2() -> dict[str, Any]:
    workbook = load_workbook(S2_SOURCE)
    readme = workbook["README"]
    target_row = None
    for row in range(2, readme.max_row + 1):
        if readme.cell(row, 1).value == "Seven required checks":
            target_row = row
            break
    if target_row is None:
        raise RuntimeError("Seven-required-checks README row not found")
    source_text = readme.cell(target_row, 2).value
    if "BUSCO ≥50 unique-site" not in source_text or "three-PEV" not in source_text:
        raise RuntimeError(f"Unexpected S2 README source text: {source_text}")
    readme.cell(target_row, 2).value = S2_REQUIRED_CHECKS
    if readme["B2"].value != "Primary exact-genome-ID Google Earth Engine (GEE) analysis and sensitivity analyses.":
        raise RuntimeError(f"Unexpected S2 hierarchy text: {readme['B2'].value}")
    readme["B2"] = (
        "Primary exact-genome-identifier (ID) Google Earth Engine (GEE) analysis and "
        "sensitivity analyses."
    )
    workbook.properties.title = "Supplemental Table S2: primary retained GEE results and verified annotations"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(S2_OUTPUT)
    return {
        "readme_cell": f"README!B{target_row}",
        "source_text": source_text,
        "output_text": S2_REQUIRED_CHECKS,
        "busco_threshold": "≥50%",
        "pev_first_expansion": "phylogenetic eigenvector (PEV)",
        "id_first_expansion": "genome-identifier (ID)",
    }


def build_s3() -> dict[str, Any]:
    workbook = load_workbook(S3_SOURCE)
    readme = workbook["README"]
    readme_changes = {
        "B2": (
            "Secondary AlphaEarth Foundations (AEF) analyses and descriptive site-level "
            "cross-representation alignment/crosswalk with named Google Earth Engine (GEE) "
            "variables; axis meanings remain distributed and nonunique."
        ),
        "B3": (
            "The 126 exact-ID genome records mapped to 90 unique coordinate sites; both "
            "representations use the same recorded coordinates, and site-level analyses give "
            "each coordinate one observation."
        ),
        "B5": "24 pooled Pfam–AEF pairs met global Benjamini–Hochberg (BH) q < 0.05.",
        "B12": (
            "All 832 descriptive site-level cross-representation correlations between 64 unitless "
            "AEF axes and 13 named GEE variables; axis meanings remain distributed and nonunique; "
            "224 met global BH q < 0.05 and 44 met Bonferroni p < 0.05."
        ),
        "B13": (
            "Named-GEE-variable summary of the 832-pair descriptive site-level "
            "cross-representation alignment/crosswalk; axis meanings remain distributed and nonunique."
        ),
        "B14": (
            "AEF-axis summary of the 832-pair descriptive site-level cross-representation "
            "alignment/crosswalk; axis meanings remain distributed and nonunique."
        ),
        "B17": (
            "For S3G–I, both representations use the same recorded coordinates. Genome records "
            "sharing a coordinate had identical AEF values and identical GEE values, so one value "
            "per coordinate was retained. For S3C–D, recorded temperature was averaged across "
            "genome records at a site; 7 sites contained more than one recorded temperature."
        ),
        "B18": (
            "S3G–I report a descriptive site-level cross-representation alignment/crosswalk; axis "
            "meanings remain distributed and nonunique; both representations use the same recorded coordinates."
        ),
    }
    for coordinate, value in readme_changes.items():
        readme[coordinate] = value

    alignment = workbook["S3G_AEF_GEE_alignment"]
    if alignment.max_row != 833 or alignment.cell(1, 10).value != "interpretation_boundary":
        raise RuntimeError("Unexpected S3G alignment dimensions/schema")
    for row in range(2, alignment.max_row + 1):
        alignment.cell(row, 10).value = S3_AFFIRMATIVE_ALIGNMENT
        alignment.cell(row, 10).alignment = Alignment(vertical="top", wrap_text=True)

    candidates = workbook["S3K_AEF_candidate_checks"]
    if candidates.max_row != 69 or candidates.cell(1, 15).value != "interpretation_boundary":
        raise RuntimeError("Unexpected S3K candidate dimensions/schema")
    for row in range(2, candidates.max_row + 1):
        candidates.cell(row, 15).value = S3_AFFIRMATIVE_CANDIDATE
        candidates.cell(row, 15).alignment = Alignment(vertical="top", wrap_text=True)

    workbook.properties.title = "Supplemental Table S3: secondary AEF analyses and cross-representation alignment"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(S3_OUTPUT)
    return {
        "readme_cells_updated": len(readme_changes),
        "alignment_boundary_cells_updated": alignment.max_row - 1,
        "candidate_boundary_cells_updated": candidates.max_row - 1,
        "exact_id_records": 126,
        "unique_coordinate_sites": 90,
        "affirmative_alignment_statement": S3_AFFIRMATIVE_ALIGNMENT,
        "affirmative_candidate_statement": S3_AFFIRMATIVE_CANDIDATE,
    }


def is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, float) and math.isnan(left):
        return isinstance(right, float) and math.isnan(right)
    return left == right


def audit_numeric_identity(source: Path, output: Path) -> dict[str, Any]:
    source_workbook = load_workbook(source, read_only=False, data_only=False)
    output_workbook = load_workbook(output, read_only=False, data_only=False)
    if source_workbook.sheetnames != output_workbook.sheetnames:
        raise RuntimeError("Workbook sheet order changed")
    numeric_compared = 0
    differences = []
    for title in source_workbook.sheetnames:
        source_sheet = source_workbook[title]
        output_sheet = output_workbook[title]
        for row in source_sheet.iter_rows():
            for source_cell in row:
                if not is_numeric(source_cell.value):
                    continue
                numeric_compared += 1
                output_value = output_sheet[source_cell.coordinate].value
                if not values_equal(source_cell.value, output_value):
                    differences.append(
                        {
                            "sheet": title,
                            "cell": source_cell.coordinate,
                            "source": source_cell.value,
                            "output": output_value,
                        }
                    )
    source_workbook.close()
    output_workbook.close()
    if differences:
        raise RuntimeError(f"Numeric cells changed: {differences[:10]}")
    return {
        "result": "PASS",
        "numeric_cells_compared": numeric_compared,
        "differences": 0,
    }


def audit_text_changes(source: Path, output: Path) -> list[dict[str, Any]]:
    source_workbook = load_workbook(source, read_only=False, data_only=False)
    output_workbook = load_workbook(output, read_only=False, data_only=False)
    changes = []
    for title in source_workbook.sheetnames:
        source_sheet = source_workbook[title]
        output_sheet = output_workbook[title]
        for row in source_sheet.iter_rows():
            for source_cell in row:
                output_value = output_sheet[source_cell.coordinate].value
                if not values_equal(source_cell.value, output_value):
                    changes.append(
                        {
                            "sheet": title,
                            "cell": source_cell.coordinate,
                            "source": source_cell.value,
                            "output": output_value,
                        }
                    )
    source_workbook.close()
    output_workbook.close()
    return changes


def audit_zip(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        if archive.testzip() is not None:
            raise RuntimeError(f"ZIP CRC test failed: {path}")


def audit_s2(change_record: dict[str, Any]) -> dict[str, Any]:
    audit_zip(S2_OUTPUT)
    numeric = audit_numeric_identity(S2_SOURCE, S2_OUTPUT)
    changes = audit_text_changes(S2_SOURCE, S2_OUTPUT)
    if len(changes) != 2 or any(change["sheet"] != "README" for change in changes):
        raise RuntimeError(f"Unexpected S2 text changes: {changes}")
    changed_outputs = {change["output"] for change in changes}
    if S2_REQUIRED_CHECKS not in changed_outputs or not any("genome-identifier (ID)" in value for value in changed_outputs):
        raise RuntimeError(f"Required S2 README changes are incomplete: {changes}")
    workbook = load_workbook(S2_OUTPUT, read_only=True, data_only=False)
    uppercase_pev_cells = []
    busco_threshold_cells = []
    bad_busco_cells = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if "PEV" in cell.value:
                    uppercase_pev_cells.append((sheet.title, cell.coordinate, cell.value))
                if "(BUSCO) ≥50%" in cell.value:
                    busco_threshold_cells.append((sheet.title, cell.coordinate))
                if "(BUSCO) ≥50 " in cell.value:
                    bad_busco_cells.append((sheet.title, cell.coordinate))
    workbook.close()
    if len(uppercase_pev_cells) != 1 or "phylogenetic eigenvector (PEV)" not in uppercase_pev_cells[0][2]:
        raise RuntimeError(f"PEV expansion audit failed: {uppercase_pev_cells}")
    if len(busco_threshold_cells) != 1 or bad_busco_cells:
        raise RuntimeError("BUSCO percentage audit failed")
    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "numeric_cell_audit": numeric,
        "text_cells_changed": len(changes),
        "declared_change": change_record,
        "uppercase_pev_occurrences": len(uppercase_pev_cells),
        "busco_ge50_percent_occurrences": len(busco_threshold_cells),
        "busco_ge50_without_percent_occurrences": len(bad_busco_cells),
    }


def audit_s3(change_record: dict[str, Any]) -> dict[str, Any]:
    audit_zip(S3_OUTPUT)
    numeric = audit_numeric_identity(S3_SOURCE, S3_OUTPUT)
    changes = audit_text_changes(S3_SOURCE, S3_OUTPUT)
    expected_change_count = 8 + 832 + 68
    if len(changes) != expected_change_count:
        raise RuntimeError(f"Unexpected S3 text-change count: {len(changes)}")

    forbidden_patterns = {
        "decoder_or_decoding": re.compile(r"\bdecod(?:er|ing)\b", re.IGNORECASE),
        "semantic": re.compile(r"\bsemantic\b", re.IGNORECASE),
        "independent_validation": re.compile(r"\bindependent validation\b", re.IGNORECASE),
        "external": re.compile(r"\bexternal\b", re.IGNORECASE),
        "internal": re.compile(r"\binternal\b", re.IGNORECASE),
        "unique_physical_label": re.compile(r"\bunique physical label", re.IGNORECASE),
        "no_physical_meaning": re.compile(r"\bno physical meaning", re.IGNORECASE),
        "does_not_assign": re.compile(r"\bdoes not assign", re.IGNORECASE),
    }
    forbidden_hits: dict[str, list[str]] = {key: [] for key in forbidden_patterns}
    component_counts = {
        "cross_representation_alignment_or_crosswalk": 0,
        "distributed_and_nonunique": 0,
        "same_recorded_coordinates": 0,
        "identical_aef_and_gee_values": 0,
        "ids_to_sites": 0,
    }
    alignment_values: set[str] = set()
    candidate_values: set[str] = set()
    workbook = load_workbook(S3_OUTPUT, read_only=True, data_only=False)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                for key, pattern in forbidden_patterns.items():
                    if pattern.search(value):
                        forbidden_hits[key].append(f"{sheet.title}!{cell.coordinate}")
                lower = value.lower()
                component_counts["cross_representation_alignment_or_crosswalk"] += (
                    "cross-representation alignment" in lower or "cross-representation" in lower and "crosswalk" in lower
                )
                component_counts["distributed_and_nonunique"] += "distributed and nonunique" in lower
                component_counts["same_recorded_coordinates"] += "same recorded coordinates" in lower
                component_counts["identical_aef_and_gee_values"] += (
                    "identical aef values and identical gee values" in lower
                )
                component_counts["ids_to_sites"] += "126 exact-id" in lower and "90 unique coordinate sites" in lower
                if sheet.title == "S3G_AEF_GEE_alignment" and cell.column == 10 and cell.row > 1:
                    alignment_values.add(value)
                if sheet.title == "S3K_AEF_candidate_checks" and cell.column == 15 and cell.row > 1:
                    candidate_values.add(value)
    workbook.close()
    populated_forbidden = {key: hits for key, hits in forbidden_hits.items() if hits}
    if populated_forbidden:
        raise RuntimeError(f"Forbidden S3 terminology remains: {populated_forbidden}")
    if alignment_values != {S3_AFFIRMATIVE_ALIGNMENT}:
        raise RuntimeError(f"S3G interpretation values differ: {alignment_values}")
    if candidate_values != {S3_AFFIRMATIVE_CANDIDATE}:
        raise RuntimeError(f"S3K interpretation values differ: {candidate_values}")
    if component_counts["distributed_and_nonunique"] < 835:
        raise RuntimeError("Distributed/nonunique axis statement is incomplete")
    if component_counts["same_recorded_coordinates"] < 835:
        raise RuntimeError("Shared-coordinate statement is incomplete")
    if component_counts["identical_aef_and_gee_values"] != 1 or component_counts["ids_to_sites"] != 1:
        raise RuntimeError("Site mapping facts are incomplete or duplicated unexpectedly")
    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "numeric_cell_audit": numeric,
        "text_cells_changed": len(changes),
        "expected_text_cells_changed": expected_change_count,
        "forbidden_terminology_hits": populated_forbidden,
        "affirmative_component_counts": component_counts,
        "alignment_boundary_values": len(alignment_values),
        "candidate_boundary_values": len(candidate_values),
        "change_record": change_record,
    }


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "relative_path": relative(path),
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
    }


def write_integrity(
    path: Path,
    source: Path,
    source_integrity: Path,
    output: Path,
    audit: dict[str, Any],
) -> None:
    if path.exists():
        raise FileExistsError(path)
    record = {
        "schema": "supplemental_workbook_terminology_integrity_v1",
        "status": "PASS",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "generator": file_record(SCRIPT),
        "software": {
            "python": platform.python_version(),
            "openpyxl": openpyxl.__version__,
        },
        "data_integrity_policy": {
            "real_data_only": True,
            "synthetic_scientific_values": False,
            "randomness_used": False,
            "all_preexisting_numeric_cells_unchanged": True,
        },
        "source_workbook": file_record(source),
        "source_integrity": file_record(source_integrity),
        "audit": audit,
        "output": file_record(output),
    }
    path.write_text(json.dumps(record, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")


def main() -> int:
    authenticate_inputs()
    s2_change = build_s2()
    s3_change = build_s3()
    s2_audit = audit_s2(s2_change)
    s3_audit = audit_s3(s3_change)
    write_integrity(S2_INTEGRITY, S2_SOURCE, S2_SOURCE_INTEGRITY, S2_OUTPUT, s2_audit)
    write_integrity(S3_INTEGRITY, S3_SOURCE, S3_SOURCE_INTEGRITY, S3_OUTPUT, s3_audit)
    print(
        json.dumps(
            {
                "table_s2_v5": {
                    "workbook": str(S2_OUTPUT.resolve()),
                    "sha256": sha256(S2_OUTPUT),
                    "integrity": str(S2_INTEGRITY.resolve()),
                    "numeric_cells_unchanged": s2_audit["numeric_cell_audit"]["numeric_cells_compared"],
                },
                "table_s3_v8": {
                    "workbook": str(S3_OUTPUT.resolve()),
                    "sha256": sha256(S3_OUTPUT),
                    "integrity": str(S3_INTEGRITY.resolve()),
                    "numeric_cells_unchanged": s3_audit["numeric_cell_audit"]["numeric_cells_compared"],
                    "text_cells_changed": s3_audit["text_cells_changed"],
                },
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
