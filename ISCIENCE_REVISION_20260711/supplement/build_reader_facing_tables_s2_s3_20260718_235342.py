#!/usr/bin/env python3
"""Create reader-facing Table S2/S3 workbook versions without changing data.

Inputs are the final audited Table S2 and Table S3 workbooks named below.  The
builder edits only declared inline-string cells and, for Table S3, one worksheet
display name.  It rewrites the ZIP members directly so every unedited archive
member remains byte-identical to its source.  The post-write audit then compares
all workbook cells, formulas, and style records with openpyxl and writes a JSON
integrity record for each output.

Generated: 2026-07-18 23:53:42 Asia/Bangkok
"""

from __future__ import annotations

import hashlib
import html
import json
import os
import platform
import re
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

import openpyxl
from openpyxl import load_workbook


STAMP = "20260718_235342"
HERE = Path(__file__).resolve().parent
BUILDER = Path(__file__).resolve()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def text_digest(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str).encode(
        "utf-8"
    )
    return hashlib.sha256(encoded).hexdigest()


TABLES = {
    "table_s2": {
        "source": HERE / "Table_S2_revised_exact_id_GEE_validation_20260712_072443.xlsx",
        "source_sha256": "5ce1b5a4c71a420fcbbf66e8bcc7af366da73a02d8df3666198aca57488cac4a",
        "output": HERE
        / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V2.xlsx",
        "integrity": HERE
        / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V2_integrity.json",
        "sheet_rename": {},
        "edits": {
            "README": {
                "A2": ("Table S2 revision", "Table S2 analysis"),
                "B2": (
                    "Exact-genome-ID GEE validation of the original discovery workflow.",
                    "Table S2 reports the exact-genome-identifier (ID) Google Earth Engine (GEE) analysis.",
                ),
                "A3": ("Correction made", "One-to-one join"),
                "B3": (
                    "The submitted workflow joined environmental and Pfam tables using non-unique species names, producing a many-to-many expansion. This revised table uses all 126 canonical genome IDs as the only join key.",
                    "Environmental and protein family (Pfam) tables were joined one-to-one using 126 canonical genome IDs.",
                ),
                "B4": (
                    "This is a bounded data-integrity correction of the original analysis, not a new discovery screen.",
                    "The exact-ID genome-level raw-count discovery analysis tested environmental-Pfam associations in the 126-genome cohort.",
                ),
                "B5": (
                    "Reconstructed raw HMM Pfam counts. No BUSCO or denominator rescaling is applied in this corrected discovery table; those specifications are reported separately as post hoc sensitivities.",
                    "Reconstructed raw hidden Markov model (HMM) Pfam counts provide the primary abundance specification; post hoc denominator and Benchmarking Universal Single-Copy Orthologs (BUSCO) specifications are reported separately.",
                ),
                "B7": (
                    "Benjamini–Hochberg and Bonferroni correction were applied across the complete 87,123-pair raw-count GEE family.",
                    "Benjamini–Hochberg (BH) false discovery rate (FDR) control and Bonferroni correction were applied across the complete 87,123-pair raw-count GEE family.",
                ),
                "B8": (
                    "Authenticated extraction of 13 environmental variables for 126 exact genome IDs.",
                    "Exact-ID extraction of 13 environmental variables, including sea surface temperature (SST), for 126 genomes.",
                ),
                "B9": (
                    "Corrected result counts by environmental variable and for the full test family.",
                    "Result counts by environmental variable and for the full test family.",
                ),
                "B10": (
                    "All 84 pairs with global Benjamini–Hochberg q < 0.05.",
                    "All 84 pairs meeting global BH FDR q < 0.05.",
                ),
            },
            "Table S2D File Index": {
                "A2": (
                    "Complete corrected correlation family",
                    "Complete correlation family",
                ),
                "A4": (
                    "Corrected correlation provenance",
                    "Correlation provenance",
                ),
            },
        },
    },
    "table_s3": {
        "source": HERE / "Table_S3_corrected_AEF_20260718_224354_V4.xlsx",
        "source_sha256": "20949a695fd451d6529265d66998cea28088ad2b34761bddedda8dfa03940d90",
        "output": HERE / f"Table_S3_AEF_{STAMP}_V5.xlsx",
        "integrity": HERE / f"Table_S3_AEF_{STAMP}_V5_integrity.json",
        "sheet_rename": {
            "S3E_corrected_output_index": "S3E_AEF_output_index",
        },
        "edits": {
            "README": {
                "A1": ("Corrected Supplemental Table S3", "Supplemental Table S3"),
                "B1": (
                    "AEF association analyses",
                    "AlphaEarth Foundations (AEF) association analyses",
                ),
                "B4": (
                    "24 pairs met global Benjamini-Hochberg q < 0.05",
                    "24 pairs met the global Benjamini–Hochberg (BH) threshold (q < 0.05).",
                ),
                "B8": (
                    "Axis-level summary of recorded-metadata correlations and corrected pooled Pfam associations.",
                    "Axis-level summary of recorded-metadata correlations and pooled Pfam associations.",
                ),
                "B9": (
                    "Files and SHA-256 hashes from the corrected AEF run.",
                    "Files and SHA-256 hashes from the AEF run.",
                ),
                "A12": ("Corrected run manifest SHA-256", "Run manifest SHA-256"),
            },
        },
    },
}


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def workbook_sheet_targets(archive: zipfile.ZipFile) -> dict[str, str]:
    import xml.etree.ElementTree as ET

    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in relationships.findall(f"{{{NS_PKG_REL}}}Relationship")
    }
    result: dict[str, str] = {}
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        rel_id = sheet.attrib[f"{{{NS_REL}}}id"]
        target = rel_targets[rel_id].lstrip("/")
        if target.startswith("xl/"):
            result[sheet.attrib["name"]] = target
        else:
            result[sheet.attrib["name"]] = f"xl/{target}"
    return result


def replace_inline_cell(
    xml_bytes: bytes,
    coordinate: str,
    old_text: str,
    new_text: str,
) -> bytes:
    xml_text = xml_bytes.decode("utf-8")
    pattern = re.compile(
        rf'(<c(?=[^>]*\br="{re.escape(coordinate)}")[^>]*>.*?<t(?:\s[^>]*)?>)'
        rf"(.*?)"
        rf"(</t>.*?</c>)",
        flags=re.DOTALL,
    )
    matches = list(pattern.finditer(xml_text))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected one inline-string cell {coordinate}; found {len(matches)}"
        )
    observed = html.unescape(matches[0].group(2))
    if observed != old_text:
        raise RuntimeError(
            f"Unexpected source text at {coordinate}: {observed!r} != {old_text!r}"
        )
    updated = pattern.sub(
        lambda match: f"{match.group(1)}{escape(new_text)}{match.group(3)}",
        xml_text,
        count=1,
    )
    return updated.encode("utf-8")


def create_output(config: dict[str, Any]) -> dict[str, Any]:
    source: Path = config["source"]
    output: Path = config["output"]
    integrity: Path = config["integrity"]
    if not source.is_file():
        raise FileNotFoundError(source)
    if sha256_file(source) != config["source_sha256"]:
        raise RuntimeError(f"Source SHA-256 changed: {source}")
    for new_path in (output, integrity):
        if new_path.exists():
            raise FileExistsError(f"Refusing to overwrite {new_path}")

    with zipfile.ZipFile(source, "r") as zin:
        source_members = {item.filename: zin.read(item.filename) for item in zin.infolist()}
        source_infos = zin.infolist()
        sheet_targets = workbook_sheet_targets(zin)

    modified_members: dict[str, bytes] = {}
    for sheet_name, sheet_edits in config["edits"].items():
        target = sheet_targets[sheet_name]
        payload = modified_members.get(target, source_members[target])
        for coordinate, (old_text, new_text) in sheet_edits.items():
            payload = replace_inline_cell(payload, coordinate, old_text, new_text)
        modified_members[target] = payload

    if config["sheet_rename"]:
        workbook_xml = source_members["xl/workbook.xml"].decode("utf-8")
        for old_name, new_name in config["sheet_rename"].items():
            expected_occurrences = 2  # sheet display name and local filter name
            observed_occurrences = workbook_xml.count(old_name)
            if observed_occurrences != expected_occurrences:
                raise RuntimeError(
                    f"Expected {expected_occurrences} workbook references to {old_name!r}; "
                    f"found {observed_occurrences}"
                )
            workbook_xml = workbook_xml.replace(old_name, new_name)
        modified_members["xl/workbook.xml"] = workbook_xml.encode("utf-8")

    output.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.name}.", suffix=".tmp", dir=output.parent
    )
    os.close(file_descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(temporary, "w") as zout:
            for info in source_infos:
                zout.writestr(info, modified_members.get(info.filename, source_members[info.filename]))
        os.replace(temporary, output)
    finally:
        if temporary.exists():
            temporary.unlink()

    audit = audit_output(source, output, config, source_members)
    audit.update(
        {
            "schema": "reader_facing_workbook_integrity_v1",
            "generated_at": datetime.now().astimezone().isoformat(),
            "builder": str(BUILDER),
            "builder_sha256": sha256_file(BUILDER),
            "python": platform.python_version(),
            "openpyxl": openpyxl.__version__,
            "source": str(source.resolve()),
            "source_sha256": sha256_file(source),
            "output": str(output.resolve()),
            "output_sha256": sha256_file(output),
            "declared_edits": {
                sheet: {
                    cell: {"old": values[0], "new": values[1]}
                    for cell, values in cells.items()
                }
                for sheet, cells in config["edits"].items()
            },
            "declared_sheet_rename": config["sheet_rename"],
        }
    )
    if audit["result"] != "PASS":
        raise RuntimeError(f"Integrity audit failed for {output}: {audit}")
    with integrity.open("x", encoding="utf-8") as handle:
        json.dump(audit, handle, indent=2, ensure_ascii=False)
        handle.write("\n")
    return audit


def serialize_cell(cell: Any) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, (datetime,)):
        value = value.isoformat()
    return {
        "value": value,
        "data_type": cell.data_type,
        "style_array": list(cell._style) if cell.has_style else None,
        "number_format": cell.number_format,
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
        "comment": cell.comment.text if cell.comment else None,
    }


def audit_output(
    source: Path,
    output: Path,
    config: dict[str, Any],
    source_members: dict[str, bytes],
) -> dict[str, Any]:
    with zipfile.ZipFile(output, "r") as zout:
        bad_member = zout.testzip()
        output_members = {item.filename: zout.read(item.filename) for item in zout.infolist()}
    changed_archive_members = sorted(
        name
        for name in source_members
        if source_members[name] != output_members.get(name)
    )
    missing_or_added_members = sorted(set(source_members) ^ set(output_members))

    source_wb = load_workbook(source, data_only=False, read_only=False)
    output_wb = load_workbook(output, data_only=False, read_only=False)
    rename = config["sheet_rename"]
    expected_output_names = [rename.get(name, name) for name in source_wb.sheetnames]

    declared_changes = {
        (sheet, coordinate): {"old": texts[0], "new": texts[1]}
        for sheet, cells in config["edits"].items()
        for coordinate, texts in cells.items()
    }
    actual_changes: list[dict[str, Any]] = []
    formula_changes: list[dict[str, Any]] = []
    style_changes: list[dict[str, Any]] = []
    undeclared_cell_changes: list[dict[str, Any]] = []
    declared_change_errors: list[dict[str, Any]] = []
    scientific_cell_changes: list[dict[str, Any]] = []
    cells_compared = 0
    formulas_compared = 0
    styles_compared = 0

    scientific_source_payload: list[Any] = []
    scientific_output_payload: list[Any] = []

    for source_sheet_name in source_wb.sheetnames:
        output_sheet_name = rename.get(source_sheet_name, source_sheet_name)
        source_ws = source_wb[source_sheet_name]
        output_ws = output_wb[output_sheet_name]
        max_row = max(source_ws.max_row, output_ws.max_row)
        max_column = max(source_ws.max_column, output_ws.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                source_cell = source_ws.cell(row, column)
                output_cell = output_ws.cell(row, column)
                coordinate = source_cell.coordinate
                cells_compared += 1
                source_serialized = serialize_cell(source_cell)
                output_serialized = serialize_cell(output_cell)
                source_style = (
                    source_serialized["style_array"],
                    source_serialized["number_format"],
                )
                output_style = (
                    output_serialized["style_array"],
                    output_serialized["number_format"],
                )
                styles_compared += 1
                if source_style != output_style:
                    style_changes.append(
                        {
                            "sheet": source_sheet_name,
                            "cell": coordinate,
                            "source": source_style,
                            "output": output_style,
                        }
                    )
                if source_cell.data_type == "f" or output_cell.data_type == "f":
                    formulas_compared += 1
                    if (
                        source_cell.data_type,
                        source_cell.value,
                    ) != (
                        output_cell.data_type,
                        output_cell.value,
                    ):
                        formula_changes.append(
                            {
                                "sheet": source_sheet_name,
                                "cell": coordinate,
                                "source": source_cell.value,
                                "output": output_cell.value,
                            }
                        )
                if source_serialized != output_serialized:
                    change = {
                        "sheet": source_sheet_name,
                        "output_sheet": output_sheet_name,
                        "cell": coordinate,
                        "source_value": source_cell.value,
                        "output_value": output_cell.value,
                        "source_type": source_cell.data_type,
                        "output_type": output_cell.data_type,
                    }
                    actual_changes.append(change)
                    expected = declared_changes.get((source_sheet_name, coordinate))
                    if expected is None:
                        undeclared_cell_changes.append(change)
                    elif (
                        source_cell.value != expected["old"]
                        or output_cell.value != expected["new"]
                    ):
                        declared_change_errors.append(change)

                is_reader_label = (
                    source_sheet_name == "README"
                    or (source_sheet_name, coordinate) in declared_changes
                )
                if not is_reader_label:
                    scientific_source_payload.append(
                        (source_sheet_name, coordinate, source_serialized)
                    )
                    scientific_output_payload.append(
                        (source_sheet_name, coordinate, output_serialized)
                    )
                    if source_serialized != output_serialized:
                        scientific_cell_changes.append(change)

    actual_change_keys = {(item["sheet"], item["cell"]) for item in actual_changes}
    missing_declared_changes = sorted(
        f"{sheet}!{cell}"
        for sheet, cell in set(declared_changes) - actual_change_keys
    )
    expected_changed_members = {
        "xl/workbook.xml" if config["sheet_rename"] else None,
    }
    expected_changed_members.discard(None)
    with zipfile.ZipFile(source, "r") as zin:
        sheet_targets = workbook_sheet_targets(zin)
    expected_changed_members.update(
        sheet_targets[sheet_name] for sheet_name in config["edits"]
    )

    checks = {
        "zip_crc_test": bad_member is None,
        "archive_member_set_unchanged": not missing_or_added_members,
        "only_expected_archive_members_changed": set(changed_archive_members)
        == expected_changed_members,
        "sheet_names_match_declared_rename": output_wb.sheetnames
        == expected_output_names,
        "every_declared_cell_changed_once": not missing_declared_changes,
        "no_undeclared_cell_changes": not undeclared_cell_changes,
        "declared_cell_values_match": not declared_change_errors,
        "no_formula_changes": not formula_changes,
        "no_style_changes": not style_changes,
        "no_scientific_cell_changes": not scientific_cell_changes,
        "scientific_payload_digest_equal": text_digest(scientific_source_payload)
        == text_digest(scientific_output_payload),
    }
    return {
        "result": "PASS" if all(checks.values()) else "FAIL",
        "checks": checks,
        "zip_bad_member": bad_member,
        "archive_members_changed": changed_archive_members,
        "archive_members_missing_or_added": missing_or_added_members,
        "source_sheet_names": source_wb.sheetnames,
        "output_sheet_names": output_wb.sheetnames,
        "cells_compared": cells_compared,
        "formulas_compared": formulas_compared,
        "styles_compared": styles_compared,
        "actual_cell_changes": actual_changes,
        "missing_declared_changes": missing_declared_changes,
        "undeclared_cell_changes": undeclared_cell_changes,
        "declared_change_errors": declared_change_errors,
        "formula_changes": formula_changes,
        "style_changes": style_changes,
        "scientific_cell_changes": scientific_cell_changes,
        "scientific_source_payload_sha256": text_digest(scientific_source_payload),
        "scientific_output_payload_sha256": text_digest(scientific_output_payload),
    }


def main() -> int:
    completed = []
    for key in ("table_s2", "table_s3"):
        audit = create_output(TABLES[key])
        completed.append(
            {
                "table": key,
                "output": audit["output"],
                "integrity": str(TABLES[key]["integrity"].resolve()),
                "result": audit["result"],
                "output_sha256": audit["output_sha256"],
                "cells_compared": audit["cells_compared"],
            }
        )
    print(json.dumps(completed, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
