#!/usr/bin/env python3
"""Build corrected Supplemental Table S3 from the validated AEF run.

All numerical results are read from the timestamped corrected real-data run or
computed from the exact 126-genome manifest/AEF join. Existing workbooks are
never overwritten.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import scipy
import statsmodels
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats
from statsmodels.stats.multitest import multipletests


SCRIPT_VERSION = "2026-07-18.2"
AXES = [f"A{i:02d}" for i in range(64)]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--aef-run", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--aef-embeddings", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--integrity-output", type=Path, required=True)
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def as_bool(series: pd.Series, name: str) -> pd.Series:
    values = series.astype(str).str.strip().str.lower()
    if not values.isin(["true", "false"]).all():
        raise ValueError(f"Unrecognized Boolean value in {name}")
    return values.eq("true")


def require_file(path: Path) -> Path:
    path = path.resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    return path


def load_exact_metadata(
    manifest_path: Path, embeddings_path: Path
) -> tuple[pd.DataFrame, pd.DataFrame]:
    manifest = pd.read_csv(manifest_path, low_memory=False)
    manifest = manifest.loc[
        as_bool(manifest["safe_for_aef_pfam_analysis"], "safe_for_aef_pfam_analysis")
    ].sort_values("master_row", kind="stable")
    aef = pd.read_csv(embeddings_path, low_memory=False)
    for frame, label in ((manifest, "manifest"), (aef, "AEF embeddings")):
        if frame["Genome"].isna().any() or frame["Genome"].duplicated().any():
            raise ValueError(f"{label} has blank or duplicate Genome identifiers")
    if len(manifest) != 126 or set(manifest["Genome"]) != set(aef["Genome"]):
        raise ValueError("The exact AEF analysis cohort is not 126 matching Genome IDs")
    aef = aef.set_index("Genome").loc[manifest["Genome"]].reset_index()
    if aef["Genome"].tolist() != manifest["Genome"].tolist():
        raise AssertionError("Exact-ID ordering failed")
    observed_axes = [column for column in aef.columns if len(column) == 3 and column[0] == "A"]
    if observed_axes != AXES:
        raise ValueError(f"Expected A00--A63; observed {observed_axes}")
    required = ["Temperature (°C)", "DD latitude", "DD longitude"]
    if manifest[required].isna().any().any() or aef[AXES].isna().any().any():
        raise ValueError("Missing values in the exact metadata screen")
    return manifest.reset_index(drop=True), aef.reset_index(drop=True)


def metadata_screen(manifest: pd.DataFrame, aef: pd.DataFrame) -> pd.DataFrame:
    variables = [
        ("Recorded temperature (°C)", manifest["Temperature (°C)"].to_numpy(float)),
        ("Recorded latitude", manifest["DD latitude"].to_numpy(float)),
        ("Recorded longitude", manifest["DD longitude"].to_numpy(float)),
    ]
    rows: list[dict[str, object]] = []
    for variable_name, values in variables:
        for axis in AXES:
            result = stats.spearmanr(aef[axis].to_numpy(float), values)
            rows.append(
                {
                    "recorded_variable": variable_name,
                    "embedding_dim": axis,
                    "n_genomes": len(values),
                    "spearman_rho": result.statistic,
                    "p_value_two_sided": result.pvalue,
                }
            )
    frame = pd.DataFrame(rows)
    frame["q_value_bh_global_192"] = multipletests(
        frame["p_value_two_sided"].to_numpy(float), method="fdr_bh"
    )[1]
    return frame


def add_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(frame.columns.tolist())
    for row in frame.itertuples(index=False, name=None):
        sheet.append(list(row))
    style_sheet(sheet)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000E+00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, cells in enumerate(sheet.iter_cols(), start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in cells[:200]]
        width = min(max(max(map(len, values), default=0) + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def main() -> None:
    args = parse_args()
    aef_run = args.aef_run.resolve()
    if not aef_run.is_dir():
        raise NotADirectoryError(aef_run)
    manifest_path = require_file(args.manifest)
    embeddings_path = require_file(args.aef_embeddings)
    output = args.output.resolve()
    integrity_output = args.integrity_output.resolve()
    for path in (output, integrity_output):
        if path.exists():
            raise FileExistsError(f"Refusing to overwrite {path}")
        path.parent.mkdir(parents=True, exist_ok=True)

    run_manifest_path = require_file(aef_run / "run_manifest.json")
    run_manifest = json.loads(run_manifest_path.read_text())
    expected_summary = run_manifest["computed_summary"]
    expected_tests = (
        int(expected_summary["strict_variable_pfam_profiles"])
        * int(expected_summary["aef_axes"])
    )
    if expected_summary["cohort_size"] != 126 or expected_summary["global_tests"] != expected_tests:
        raise ValueError("Unexpected corrected AEF run dimensions")

    pooled_sig_path = require_file(aef_run / "pooled_126_significant_q_lt_0.05.csv")
    pooled_full_path = require_file(aef_run / "pooled_126_full_correlations.csv.gz")
    pooled_dim_path = require_file(aef_run / "pooled_126_dimension_summary.csv")
    centered_sig_path = require_file(
        aef_run / "phylum_centered_126_significant_q_lt_0.05.csv"
    )
    within_summary_path = require_file(aef_run / "within_phylum_specification_summary.csv")

    pooled_sig = pd.read_csv(pooled_sig_path).sort_values(
        ["q_value_bh_global", "p_value_two_sided", "pfam", "embedding_dim"],
        kind="stable",
    )
    pooled_full = pd.read_csv(pooled_full_path)
    pooled_dim = pd.read_csv(pooled_dim_path)
    centered_sig = pd.read_csv(centered_sig_path).sort_values(
        ["q_value_bh_global", "p_value_two_sided", "pfam", "embedding_dim"],
        kind="stable",
    )
    within_summary = pd.read_csv(within_summary_path)
    if len(pooled_sig) != expected_summary["bh_q_lt_0.05"]:
        raise ValueError("Pooled significant-row count differs from run manifest")
    if len(centered_sig) != expected_summary["phylum_centered_screen"]["n_q_lt_0.05"]:
        raise ValueError("Phylum-centered row count differs from run manifest")

    manifest, aef = load_exact_metadata(manifest_path, embeddings_path)
    metadata = metadata_screen(manifest, aef)

    best_pfam = (
        pooled_full.sort_values(
            ["embedding_dim", "q_value_bh_global", "p_value_two_sided", "pfam"],
            kind="stable",
        )
        .groupby("embedding_dim", sort=False)
        .first()
        .reset_index()
    )
    metadata_wide = metadata.pivot(
        index="embedding_dim", columns="recorded_variable", values="spearman_rho"
    )
    dimension_rows: list[dict[str, object]] = []
    dim_lookup = pooled_dim.set_index("embedding_dim")
    best_lookup = best_pfam.set_index("embedding_dim")
    for axis in AXES:
        axis_meta = metadata.loc[metadata["embedding_dim"].eq(axis)].copy()
        strongest = axis_meta.iloc[axis_meta["spearman_rho"].abs().argmax()]
        dimension_rows.append(
            {
                "embedding_dim": axis,
                "recorded_temperature_rho": metadata_wide.loc[
                    axis, "Recorded temperature (°C)"
                ],
                "recorded_latitude_rho": metadata_wide.loc[axis, "Recorded latitude"],
                "recorded_longitude_rho": metadata_wide.loc[axis, "Recorded longitude"],
                "strongest_recorded_variable": strongest["recorded_variable"],
                "strongest_recorded_rho": strongest["spearman_rho"],
                "n_pfam_pairs_global_bh_q_lt_0.05": int(
                    dim_lookup.loc[axis, "n_q_lt_0.05"]
                ),
                "strongest_pfam": best_lookup.loc[axis, "pfam"],
                "strongest_pfam_rho": best_lookup.loc[axis, "spearman_rho"],
                "strongest_pfam_global_bh_q": best_lookup.loc[
                    axis, "q_value_bh_global"
                ],
                "n_recorded_metadata_pairs_global_bh_q_lt_0.05": int(
                    (axis_meta["q_value_bh_global_192"] < 0.05).sum()
                ),
            }
        )
    dimension_summary = pd.DataFrame(dimension_rows)

    index_rows: list[dict[str, object]] = []
    output_hashes = run_manifest.get("output_files_sha256", {})
    for filename, digest in sorted(output_hashes.items()):
        path = aef_run / filename
        if not path.is_file() or sha256(path) != digest:
            raise ValueError(f"Corrected-run output hash mismatch: {filename}")
        index_rows.append(
            {
                "file": filename,
                "sha256": digest,
                "bytes": path.stat().st_size,
            }
        )
    output_index = pd.DataFrame(index_rows)

    workbook = Workbook()
    workbook.remove(workbook.active)
    readme = workbook.create_sheet("README")
    pooled_pair_count = len(pooled_sig)
    centered_pair_count = len(centered_sig)
    pfam_profile_count = int(expected_summary["strict_variable_pfam_profiles"])
    axis_count = int(expected_summary["aef_axes"])
    readme_rows = [
        ("Corrected Supplemental Table S3", "AEF association analyses"),
        ("Analysis cohort", "126 exact-ID macroalgal genomes"),
        (
            "Pooled family",
            f"{pfam_profile_count:,} strict Pfam accessions x {axis_count} AEF axes = {expected_tests:,} tests",
        ),
        (
            "Pooled result",
            f"{pooled_pair_count:,} pairs met global Benjamini-Hochberg q < 0.05",
        ),
        (
            "S3A",
            f"All {pooled_pair_count:,} pooled raw-count Pfam-AEF pairs meeting global BH q < 0.05.",
        ),
        (
            "S3B",
            f"All {centered_pair_count:,} pairs meeting global BH q < 0.05 after subtracting each phylum-specific Pfam mean.",
        ),
        (
            "S3C",
            "All 192 Spearman correlations of A00-A63 with recorded temperature, latitude, and longitude; BH correction spans 192 tests.",
        ),
        (
            "S3D",
            "Axis-level summary of recorded-metadata correlations and corrected pooled Pfam associations.",
        ),
        ("S3E", "Files and SHA-256 hashes from the corrected AEF run."),
        (
            "S3F",
            "Within-phylum screens using the stated nonzero-prevalence specifications.",
        ),
        ("AEF interpretation", "A00-A63 are unitless latent geospatial axes."),
        (
            "Corrected run manifest SHA-256",
            sha256(run_manifest_path),
        ),
    ]
    for row in readme_rows:
        readme.append(row)
    style_sheet(readme)
    readme.auto_filter.ref = "A1:B1"

    add_dataframe_sheet(workbook, "S3A_pooled_AEF_Pfam", pooled_sig.reset_index(drop=True))
    add_dataframe_sheet(
        workbook, "S3B_phylum_centered", centered_sig.reset_index(drop=True)
    )
    add_dataframe_sheet(
        workbook, "S3C_AEF_recorded_metadata", metadata.reset_index(drop=True)
    )
    add_dataframe_sheet(
        workbook, "S3D_AEF_axis_summary", dimension_summary.reset_index(drop=True)
    )
    add_dataframe_sheet(workbook, "S3E_corrected_output_index", output_index)
    add_dataframe_sheet(workbook, "S3F_within_phylum_summary", within_summary)

    workbook.properties.title = "Corrected Supplemental Table S3: AEF analyses"
    workbook.properties.creator = "Reproducible revision pipeline"
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(output)

    check = load_workbook(output, read_only=True, data_only=True)
    expected_sheets = [
        "README",
        "S3A_pooled_AEF_Pfam",
        "S3B_phylum_centered",
        "S3C_AEF_recorded_metadata",
        "S3D_AEF_axis_summary",
        "S3E_corrected_output_index",
        "S3F_within_phylum_summary",
    ]
    if check.sheetnames != expected_sheets:
        raise AssertionError(check.sheetnames)
    row_counts = {sheet.title: sheet.max_row - 1 for sheet in check.worksheets}
    expected_counts = {
        "S3A_pooled_AEF_Pfam": len(pooled_sig),
        "S3B_phylum_centered": len(centered_sig),
        "S3C_AEF_recorded_metadata": len(metadata),
        "S3D_AEF_axis_summary": 64,
        "S3E_corrected_output_index": len(output_index),
        "S3F_within_phylum_summary": len(within_summary),
    }
    for sheet, expected in expected_counts.items():
        if row_counts[sheet] != expected:
            raise AssertionError(f"{sheet}: {row_counts[sheet]} != {expected}")
    check.close()

    inputs = {
        "corrected_run_manifest": run_manifest_path,
        "reconciled_manifest": manifest_path,
        "aef_embeddings": embeddings_path,
        "pooled_significant": pooled_sig_path,
        "pooled_full": pooled_full_path,
        "pooled_dimension_summary": pooled_dim_path,
        "phylum_centered_significant": centered_sig_path,
        "within_phylum_summary": within_summary_path,
    }
    integrity = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "generator": {
            "path": str(Path(__file__).resolve()),
            "sha256": sha256(Path(__file__).resolve()),
        },
        "command": [str(value) for value in sys.argv],
        "software": {
            "python": platform.python_version(),
            "numpy": np.__version__,
            "pandas": pd.__version__,
            "scipy": scipy.__version__,
            "statsmodels": statsmodels.__version__,
            "openpyxl": openpyxl.__version__,
        },
        "inputs": {
            name: {"path": str(path), "sha256": sha256(path), "bytes": path.stat().st_size}
            for name, path in inputs.items()
        },
        "output": {
            "path": str(output),
            "sha256": sha256(output),
            "bytes": output.stat().st_size,
            "sheet_data_rows": row_counts,
        },
        "computed_counts": {
            "pooled_global_bh_pairs": len(pooled_sig),
            "phylum_centered_global_bh_pairs": len(centered_sig),
            "recorded_metadata_tests": len(metadata),
            "within_phylum_specifications": len(within_summary),
        },
    }
    integrity_output.write_text(json.dumps(integrity, indent=2, sort_keys=True) + "\n")
    print(output)
    print(integrity_output)


if __name__ == "__main__":
    main()
