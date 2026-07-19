#!/usr/bin/env python3
"""Build reader-facing Table S2 V3 and Table S3 V6 from audited real-data outputs.

Table S2 keeps the exact-ID GEE discovery tables and adds the complete selected-
84 non-null sensitivity results, the independently audited 99,999-permutation
structured-null refinement, and the corresponding seven-check candidate table.

Table S3 keeps the pooled and phylum-centered AEF results, replaces the earlier
genome-weighted recorded-metadata summaries with unique-coordinate-site
summaries, and adds the complete site-level AEF--GEE alignment plus the corrected
99,999-permutation selected-AEF structured-null and candidate tables.

Every numerical value is read from an authenticated source file or computed from
the authenticated 126-genome exact-ID AEF/metadata join. No synthetic or
placeholder scientific data are generated. Existing workbooks are not changed.

Created: 2026-07-19 21:06:11 Asia/Bangkok
"""

from __future__ import annotations

import hashlib
import json
import platform
import sys
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import openpyxl
import pandas as pd
import scipy
import statsmodels
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats
from statsmodels.stats.multitest import multipletests


SCRIPT_VERSION = "2026-07-19.1"
STAMP = "20260719_211551"
SCRIPT = Path(__file__).resolve()
SUPPLEMENT = SCRIPT.parent
ROOT = SCRIPT.parents[2]

S2_SOURCE = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260718_235342_V2.xlsx"
S3_SOURCE = SUPPLEMENT / "Table_S3_AEF_20260718_235342_V5.xlsx"
S2_OUTPUT = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V3.xlsx"
S3_OUTPUT = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V6.xlsx"
S2_INTEGRITY = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V3_integrity.json"
S3_INTEGRITY = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V6_integrity.json"

GEE_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "gee_validation"
ANALYSIS_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "analysis_stats"
INTEGRITY_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "integrity"
AEF_RUN_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "aef" / "full_aef_corrected_run_20260718_154213"

GEE_SENSITIVITY = GEE_DIR / "GEE_primary_selected84_sensitivity_20260719_205317.csv"
GEE_BASE_MANIFEST = GEE_DIR / "GEE_primary_selected84_manifest_20260719_205317.json"
GEE_REFINED_NULL = GEE_DIR / "GEE_primary_selected84_structured_null_refined99999_20260719_210049.csv"
GEE_REFINED_CANDIDATES = GEE_DIR / "GEE_primary_selected84_candidate_summary_refined99999_20260719_210049.csv"
GEE_REFINED_MANIFEST = GEE_DIR / "GEE_primary_selected84_manifest_refined99999_20260719_210049.json"
GEE_DISCOVERY_FULL = GEE_DIR / "exact_id_gee_raw_pfam_correlations_20260712_072151.csv.gz"
GEE_EXTRACTION_MANIFEST = GEE_DIR / "exact_id_gee_environmental_extraction_manifest_20260712_071838.json"
GEE_DISCOVERY_MANIFEST = GEE_DIR / "exact_id_gee_correlation_validation_manifest_20260712_072151.json"

AEF_ALIGNMENT = GEE_DIR / "AEF_GEE_site_alignment_832pairs_20260719_204925.csv"
AEF_ALIGNMENT_VARIABLES = GEE_DIR / "AEF_GEE_site_alignment_variable_summary_20260719_204925.csv"
AEF_ALIGNMENT_AXES = GEE_DIR / "AEF_GEE_site_alignment_axis_summary_20260719_204925.csv"
AEF_ALIGNMENT_MANIFEST = GEE_DIR / "AEF_GEE_site_alignment_manifest_20260719_204925.json"
AEF_CORRECTED_NULL = ANALYSIS_DIR / "corrected_AEF_structured_null_conditional_tail_20260719_210254.csv"
AEF_CORRECTED_CANDIDATES = ANALYSIS_DIR / "corrected_AEF_robust_candidate_table_20260719_210254.csv"
AEF_CORRECTED_MANIFEST = ANALYSIS_DIR / "corrected_AEF_structured_null_manifest_20260719_210254.json"
AEF_CORRECTION_SCRIPT = ANALYSIS_DIR / "correct_aef_structured_null_conditional_tail_20260719_205621.py"

RECONCILED_MANIFEST = INTEGRITY_DIR / "reconciled_analysis_manifest_20260711_110650.csv"
AEF_EMBEDDINGS = ROOT / "AlphaEarth" / "CSV" / "alphaearth_embeddings_20251019_122918.csv"
AEF_RUN_MANIFEST = AEF_RUN_DIR / "run_manifest.json"

EXPECTED_FIXED_HASHES = {
    S2_SOURCE: "652c65d2b84f15ad4223a4b1050a769720249a84e12a790e763fef82224979a8",
    S3_SOURCE: "b350c73a1a72e56182b186903391533f5a10b9f9014aeeea0d4f3e06e130fff5",
    GEE_SENSITIVITY: "4cf5588254756d3cc050fab62f0719a941a431b72e6ee47fe8eee461b5478d15",
    GEE_REFINED_NULL: "f3877b6d79cb4eb9a698fca7be0ed995f7772860ed1329f2dcf532470c1e7964",
    GEE_REFINED_CANDIDATES: "85bd71ff3128219949e74b168a8714f9b65f2020b5854a0ee4ff0f05059d2e76",
    AEF_ALIGNMENT: "9c0c49769b751c18a88b1c1d6911f7f03749b29e34de37593948494a4464192e",
    AEF_ALIGNMENT_VARIABLES: "f3944d319c1ead505fc376a8714694a286867b91914457d42347f91134850a46",
    AEF_ALIGNMENT_AXES: "65da048c554c28dd9473046b3a78cc2fc9b07509339f48825e45f632193b673c",
    AEF_CORRECTED_NULL: "8e3e3b04193e8c1ef82ee306ab6341d4a76caeede7ded17cced1c34064544a30",
    AEF_CORRECTED_CANDIDATES: "77c55be1e1b735c6dc81081d52bafe9aed023b69d553e0f373f0084589989829",
    AEF_CORRECTED_MANIFEST: "dc44921c05dce92599e792d1d2d1b0ccaeff480f4fc09657590e92e34e2fb422",
    RECONCILED_MANIFEST: "5880f930192d4cbb11a7563825c853e458fb1690b01c9cf7cc83323e7541bd67",
    AEF_EMBEDDINGS: "e0b05e727aec4a5b45565c9026de21c46561a18c06786f235eae371771a4cf87",
}

AXES = [f"A{index:02d}" for index in range(64)]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT))


def authenticate_fixed_inputs() -> None:
    for path, expected in EXPECTED_FIXED_HASHES.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        observed = sha256(path)
        if observed != expected:
            raise RuntimeError(f"Authenticated input changed: {path}: {observed}")
    for path in (
        GEE_BASE_MANIFEST,
        GEE_REFINED_MANIFEST,
        GEE_DISCOVERY_FULL,
        GEE_EXTRACTION_MANIFEST,
        GEE_DISCOVERY_MANIFEST,
        AEF_ALIGNMENT_MANIFEST,
        AEF_CORRECTION_SCRIPT,
        AEF_RUN_MANIFEST,
    ):
        if not path.is_file():
            raise FileNotFoundError(path)


def verify_manifest_output(manifest_path: Path, output_path: Path) -> None:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    expected: str | None = None
    for record in manifest.get("outputs", {}).values():
        if not isinstance(record, dict):
            continue
        candidate = record.get("realpath") or record.get("path")
        if candidate and Path(candidate).name == output_path.name:
            expected = record.get("sha256")
            break
    if expected is None:
        raise RuntimeError(f"Manifest does not identify {output_path.name}: {manifest_path}")
    if expected != sha256(output_path):
        raise RuntimeError(f"Manifest output hash mismatch: {output_path}")


def read_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    values = sheet.values
    columns = next(values)
    frame = pd.DataFrame(values, columns=columns)
    workbook.close()
    return frame


def python_scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def style_sheet(sheet, wrap_columns: Iterable[str] = ()) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    wrap_columns = set(wrap_columns)
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 32
    headers = {cell.column: cell.value for cell in sheet[1]}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                vertical="top", wrap_text=headers.get(cell.column) in wrap_columns
            )
            if isinstance(cell.value, float):
                cell.number_format = "0.000000E+00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, cells in enumerate(sheet.iter_cols(), start=1):
        header = headers.get(column_index)
        values = [str(cell.value) if cell.value is not None else "" for cell in cells[:500]]
        cap = 55 if header in wrap_columns else 38
        width = min(max(max(map(len, values), default=0) + 2, 10), cap)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def add_frame_sheet(workbook, title: str, frame: pd.DataFrame, index: int | None = None) -> None:
    if title in workbook.sheetnames:
        raise RuntimeError(f"Sheet already exists: {title}")
    sheet = workbook.create_sheet(title, index) if index is not None else workbook.create_sheet(title)
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append([python_scalar(value) for value in row])
    wrap_columns = {
        column
        for column in frame.columns
        if any(token in str(column).lower() for token in ("statement", "boundary", "status", "role", "path", "method", "criterion", "name", "url"))
    }
    style_sheet(sheet, wrap_columns)


def replace_sheet_with_frame(workbook, title: str, frame: pd.DataFrame, index: int) -> None:
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    add_frame_sheet(workbook, title, frame, index=index)


def file_record(role: str, path: Path, rows_or_status: str) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(path)
    return {
        "file_role": role,
        "relative_path": relative(path),
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
        "rows_or_status": rows_or_status,
    }


def validate_gee_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    verify_manifest_output(GEE_BASE_MANIFEST, GEE_SENSITIVITY)
    verify_manifest_output(GEE_REFINED_MANIFEST, GEE_REFINED_NULL)
    verify_manifest_output(GEE_REFINED_MANIFEST, GEE_REFINED_CANDIDATES)
    sensitivity = pd.read_csv(GEE_SENSITIVITY, low_memory=False)
    structured_null = pd.read_csv(GEE_REFINED_NULL, low_memory=False)
    candidates = pd.read_csv(GEE_REFINED_CANDIDATES, low_memory=False)
    pair_columns = ["pfam", "environment"]
    selected_pairs = candidates[pair_columns].drop_duplicates()
    if len(sensitivity) != 2016 or len(structured_null) != 84 or len(candidates) != 84:
        raise RuntimeError("Unexpected selected-GEE output dimensions")
    if sensitivity.groupby(pair_columns).size().nunique() != 1 or int(sensitivity.groupby(pair_columns).size().iloc[0]) != 24:
        raise RuntimeError("The selected-GEE sensitivity is not 24 rows per pair")
    pair_sets = [
        set(map(tuple, frame[pair_columns].drop_duplicates().to_numpy()))
        for frame in (sensitivity, structured_null, candidates)
    ]
    if len(selected_pairs) != 84 or not (pair_sets[0] == pair_sets[1] == pair_sets[2]):
        raise RuntimeError("Selected-GEE pair sets disagree")
    all_seven = int(candidates["robust_candidate_all_required_checks"].sum())
    bonf_all_seven = int(
        (
            candidates["discovery_bonferroni_significant"]
            & candidates["robust_candidate_all_required_checks"]
        ).sum()
    )
    bonf_total = int(candidates["discovery_bonferroni_significant"].sum())
    null_pass = int((structured_null["selected_GEE_empirical_bh_q"] < 0.05).sum())
    if (all_seven, bonf_all_seven, bonf_total, null_pass) != (49, 13, 13, 56):
        raise RuntimeError("Refined selected-GEE result counts changed")
    return sensitivity, structured_null, candidates, {
        "selected_pairs": 84,
        "non_null_rows": len(sensitivity),
        "non_null_rows_per_pair": 24,
        "structured_null_rows": len(structured_null),
        "permutations": int(structured_null["permutations"].unique().item()),
        "structured_null_q_lt_0.05": null_pass,
        "all_seven": all_seven,
        "bonferroni_discovery_all_seven": bonf_all_seven,
        "bonferroni_discovery_total": bonf_total,
    }


def build_s2(
    sensitivity: pd.DataFrame,
    structured_null: pd.DataFrame,
    candidates: pd.DataFrame,
    counts: dict[str, Any],
) -> tuple[dict[str, pd.DataFrame], list[str]]:
    workbook = load_workbook(S2_SOURCE)
    retained = ["Table S2A Exact-ID GEE", "Table S2B Summary", "Table S2C FDR pairs"]
    extraction = read_sheet(S2_SOURCE, "Table S2A Exact-ID GEE")
    discovery_summary = read_sheet(S2_SOURCE, "Table S2B Summary")
    discovery = read_sheet(S2_SOURCE, "Table S2C FDR pairs")
    summary_total = discovery_summary.loc[discovery_summary["variable"].eq("TOTAL")].iloc[0]
    per_variable_summary = discovery_summary.loc[~discovery_summary["variable"].eq("TOTAL")]
    discovery_tests = int(summary_total["total_tests"])
    discovery_pairs = len(discovery)
    discovery_pfams = int(discovery["pfam"].nunique())
    discovery_bonferroni = int(discovery["sig_bonferroni05"].sum())
    named_variables = len(per_variable_summary)
    eligible_pfams_min = int(per_variable_summary["total_tests"].min())
    eligible_pfams = int(per_variable_summary["total_tests"].max())
    cohort_size = len(extraction)
    for title in ("README", "Table S2D File Index"):
        workbook.remove(workbook[title])

    readme = pd.DataFrame(
        [
            ("Table S2 analysis hierarchy", "Primary exact-genome-ID Google Earth Engine (GEE) analysis and sensitivity analyses."),
            ("Primary discovery", f"The one-to-one {cohort_size}-genome screen tested {named_variables} named environmental variables, with {eligible_pfams_min:,}–{eligible_pfams:,} estimable Pfam accessions per variable ({discovery_tests:,} tests)."),
            ("Discovery results", f"{discovery_pairs} associations representing {discovery_pfams} Pfam accessions met global Benjamini–Hochberg (BH) q < 0.05; {discovery_bonferroni} met Bonferroni p < 0.05."),
            ("S2A", f"Exact-ID extraction of {named_variables} named GEE variables for {cohort_size} genomes."),
            ("S2B", "Discovery result counts by named environmental variable and across the complete test family."),
            ("S2C", f"All {discovery_pairs} raw-count Pfam–GEE pairs meeting global BH q < 0.05 in the {discovery_tests:,}-test family."),
            ("S2D", "File-level provenance, SHA-256 hashes, byte sizes, and row counts."),
            ("S2E", f"All {counts['non_null_rows']:,} non-null sensitivity rows: {counts['selected_pairs']} selected discovery pairs × 3 abundance representations × {counts['non_null_rows_per_pair'] // 3} specifications."),
            ("S2F", f"Structured-null results for the {counts['selected_pairs']} selected pairs using {counts['permutations']:,} intact-site environmental-label permutations within site phylum-composition strata and BH correction across {counts['selected_pairs']} pairs."),
            ("S2G", f"Seven-check candidate summary. {counts['all_seven']} of {counts['selected_pairs']} selected pairs passed all seven checks; all {counts['bonferroni_discovery_all_seven']} of {counts['bonferroni_discovery_total']} Bonferroni discovery pairs passed all seven."),
            ("Seven required checks", "Unique-site raw counts; unique-site total-hit normalization; unique-site peptide normalization; quality/phylum site-cluster model; BUSCO ≥50 unique-site analysis; three-PEV topology-aware site-cluster model; and the structured null."),
            ("Inference", "These tables quantify observational named-variable associations and their sensitivity across related specifications."),
        ],
        columns=["Item", "Statement"],
    )

    index_rows = [
        file_record(f"Complete {discovery_tests:,}-pair GEE discovery family", GEE_DISCOVERY_FULL, f"{discovery_tests:,} rows"),
        file_record("Exact-ID GEE extraction manifest", GEE_EXTRACTION_MANIFEST, "validated"),
        file_record("Exact-ID GEE discovery manifest", GEE_DISCOVERY_MANIFEST, "validated"),
        file_record(f"Selected-{counts['selected_pairs']} non-null sensitivity", GEE_SENSITIVITY, f"{len(sensitivity):,} rows"),
        file_record("Selected-84 non-null sensitivity manifest", GEE_BASE_MANIFEST, "validated"),
        file_record(f"Selected-{counts['selected_pairs']} refined structured null", GEE_REFINED_NULL, f"{len(structured_null)} rows"),
        file_record(f"Selected-{counts['selected_pairs']} refined candidate summary", GEE_REFINED_CANDIDATES, f"{len(candidates)} rows"),
        file_record("Selected-84 refined structured-null manifest", GEE_REFINED_MANIFEST, "validated"),
        file_record("Workbook source version", S2_SOURCE, "preserved"),
        file_record("Workbook builder", SCRIPT, "executed"),
    ]
    file_index = pd.DataFrame(index_rows)

    add_frame_sheet(workbook, "README", readme, index=0)
    add_frame_sheet(workbook, "Table S2D File Index", file_index, index=4)
    add_frame_sheet(workbook, "S2E GEE sensitivity", sensitivity)
    add_frame_sheet(workbook, "S2F Structured null", structured_null)
    add_frame_sheet(workbook, "S2G Candidate summary", candidates)
    workbook.properties.title = "Supplemental Table S2: primary exact-ID GEE analyses"
    workbook.properties.creator = "Reproducible revision pipeline"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    if S2_OUTPUT.exists():
        raise FileExistsError(S2_OUTPUT)
    workbook.save(S2_OUTPUT)
    expected_frames = {
        "README": readme,
        "Table S2D File Index": file_index,
        "S2E GEE sensitivity": sensitivity,
        "S2F Structured null": structured_null,
        "S2G Candidate summary": candidates,
    }
    return expected_frames, retained


def load_site_metadata() -> tuple[pd.DataFrame, dict[str, Any]]:
    manifest = pd.read_csv(RECONCILED_MANIFEST, low_memory=False)
    safe = manifest["safe_for_aef_pfam_analysis"].astype(str).str.strip().str.lower().eq("true")
    manifest = manifest.loc[safe].sort_values("master_row", kind="stable").reset_index(drop=True)
    embeddings = pd.read_csv(AEF_EMBEDDINGS, low_memory=False)
    if len(manifest) != 126 or manifest["Genome"].nunique() != 126:
        raise RuntimeError("Reconciled AEF cohort is not 126 exact genome IDs")
    if len(embeddings) != 126 or embeddings["Genome"].nunique() != 126:
        raise RuntimeError("AEF embedding input is not 126 exact genome IDs")
    if set(manifest["Genome"]) != set(embeddings["Genome"]):
        raise RuntimeError("Manifest and AEF embedding genome-ID sets differ")
    observed_axes = [column for column in embeddings.columns if len(column) == 3 and column.startswith("A")]
    if observed_axes != AXES:
        raise RuntimeError(f"Expected A00--A63; observed {observed_axes}")
    selected = manifest[["Genome", "DD latitude", "DD longitude", "Temperature (°C)"]]
    merged = selected.merge(
        embeddings[["Genome", "DD latitude", "DD longitude"] + AXES],
        on="Genome",
        suffixes=("_manifest", "_aef"),
        validate="one_to_one",
    )
    for coordinate in ("DD latitude", "DD longitude"):
        delta = (
            pd.to_numeric(merged[f"{coordinate}_manifest"], errors="raise")
            - pd.to_numeric(merged[f"{coordinate}_aef"], errors="raise")
        ).abs()
        if float(delta.max()) > 1e-10:
            raise RuntimeError(f"Coordinate disagreement in {coordinate}")
    merged["site_id"] = (
        merged["DD latitude_manifest"].round(10).map(lambda value: f"{value:.10f}")
        + "|"
        + merged["DD longitude_manifest"].round(10).map(lambda value: f"{value:.10f}")
    )
    grouped = merged.groupby("site_id", sort=True)
    if grouped.ngroups != 90:
        raise RuntimeError(f"Expected 90 coordinate sites; observed {grouped.ngroups}")
    aef_disagreements = grouped[AXES].nunique(dropna=False).gt(1).any(axis=1)
    if aef_disagreements.any():
        raise RuntimeError("AEF values differ within an exact coordinate site")
    temperature_unique = grouped["Temperature (°C)"].nunique(dropna=False)
    site = grouped[AXES].first()
    site["recorded_temperature_site_mean_c"] = grouped["Temperature (°C)"].mean()
    site["recorded_latitude"] = grouped["DD latitude_manifest"].first()
    site["recorded_longitude"] = grouped["DD longitude_manifest"].first()
    site["n_genomes"] = grouped.size()
    site["n_recorded_temperature_values"] = temperature_unique
    return site, {
        "exact_id_genomes": len(merged),
        "unique_coordinate_sites": len(site),
        "repeated_coordinate_sites": int((site["n_genomes"] > 1).sum()),
        "maximum_genomes_per_site": int(site["n_genomes"].max()),
        "within_site_aef_disagreements": int(aef_disagreements.sum()),
        "sites_with_multiple_recorded_temperatures": int((temperature_unique > 1).sum()),
        "recorded_temperature_site_aggregation": "arithmetic mean across exact-ID genomes at each coordinate",
    }


def site_metadata_screen(site: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    variables = [
        ("Recorded temperature site mean (°C)", "recorded_temperature_site_mean_c", "site arithmetic mean"),
        ("Recorded latitude", "recorded_latitude", "exact coordinate"),
        ("Recorded longitude", "recorded_longitude", "exact coordinate"),
    ]
    rows: list[dict[str, Any]] = []
    for display, column, aggregation in variables:
        for axis in AXES:
            pair = site[[axis, column]].dropna()
            result = stats.spearmanr(pair[axis].to_numpy(float), pair[column].to_numpy(float))
            rows.append(
                {
                    "recorded_variable": display,
                    "aef_axis": axis,
                    "n_unique_sites": len(pair),
                    "site_aggregation": aggregation,
                    "spearman_rho": float(result.statistic),
                    "p_value_two_sided": float(result.pvalue),
                }
            )
    correlations = pd.DataFrame(rows)
    if len(correlations) != 192 or correlations["p_value_two_sided"].isna().any():
        raise RuntimeError("The site-level recorded-metadata family is incomplete")
    correlations["global_bh_q_192"] = multipletests(
        correlations["p_value_two_sided"].to_numpy(float), method="fdr_bh"
    )[1]
    correlations["global_bonferroni_p_192"] = np.minimum(
        correlations["p_value_two_sided"].to_numpy(float) * 192, 1.0
    )
    correlations["global_bh_q_lt_0.05"] = correlations["global_bh_q_192"] < 0.05
    correlations["global_bonferroni_p_lt_0.05"] = correlations["global_bonferroni_p_192"] < 0.05

    old_axis = read_sheet(S3_SOURCE, "S3D_AEF_axis_summary")
    pooled_columns = [
        "embedding_dim",
        "n_pfam_pairs_global_bh_q_lt_0.05",
        "strongest_pfam",
        "strongest_pfam_rho",
        "strongest_pfam_global_bh_q",
    ]
    pooled = old_axis[pooled_columns].rename(columns={"embedding_dim": "aef_axis"})
    summary_rows: list[dict[str, Any]] = []
    for axis in AXES:
        axis_rows = correlations.loc[correlations["aef_axis"].eq(axis)].copy()
        strongest = axis_rows.iloc[axis_rows["spearman_rho"].abs().argmax()]
        lookup = axis_rows.set_index("recorded_variable")
        summary_rows.append(
            {
                "aef_axis": axis,
                "recorded_temperature_site_mean_rho": lookup.loc["Recorded temperature site mean (°C)", "spearman_rho"],
                "recorded_latitude_rho": lookup.loc["Recorded latitude", "spearman_rho"],
                "recorded_longitude_rho": lookup.loc["Recorded longitude", "spearman_rho"],
                "strongest_site_recorded_variable": strongest["recorded_variable"],
                "strongest_site_recorded_rho": strongest["spearman_rho"],
                "strongest_site_recorded_global_bh_q": strongest["global_bh_q_192"],
                "n_site_recorded_pairs_global_bh_q_lt_0.05": int(axis_rows["global_bh_q_lt_0.05"].sum()),
                "n_site_recorded_pairs_global_bonferroni_p_lt_0.05": int(axis_rows["global_bonferroni_p_lt_0.05"].sum()),
            }
        )
    axis_summary = pd.DataFrame(summary_rows).merge(pooled, on="aef_axis", validate="one_to_one")
    computed = {
        "tests": len(correlations),
        "global_bh_q_lt_0.05": int(correlations["global_bh_q_lt_0.05"].sum()),
        "global_bonferroni_p_lt_0.05": int(correlations["global_bonferroni_p_lt_0.05"].sum()),
    }
    if computed != {"tests": 192, "global_bh_q_lt_0.05": 51, "global_bonferroni_p_lt_0.05": 23}:
        raise RuntimeError(f"Site-level metadata result counts changed: {computed}")
    return correlations, axis_summary, computed


def validate_aef_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    verify_manifest_output(AEF_ALIGNMENT_MANIFEST, AEF_ALIGNMENT)
    verify_manifest_output(AEF_ALIGNMENT_MANIFEST, AEF_ALIGNMENT_VARIABLES)
    verify_manifest_output(AEF_ALIGNMENT_MANIFEST, AEF_ALIGNMENT_AXES)
    verify_manifest_output(AEF_CORRECTED_MANIFEST, AEF_CORRECTED_NULL)
    verify_manifest_output(AEF_CORRECTED_MANIFEST, AEF_CORRECTED_CANDIDATES)
    alignment = pd.read_csv(AEF_ALIGNMENT, low_memory=False)
    variable_summary = pd.read_csv(AEF_ALIGNMENT_VARIABLES, low_memory=False)
    axis_alignment = pd.read_csv(AEF_ALIGNMENT_AXES, low_memory=False)
    corrected_null = pd.read_csv(AEF_CORRECTED_NULL, low_memory=False)
    corrected_candidates = pd.read_csv(AEF_CORRECTED_CANDIDATES, low_memory=False)
    if (len(alignment), len(variable_summary), len(axis_alignment), len(corrected_null), len(corrected_candidates)) != (832, 13, 64, 204, 68):
        raise RuntimeError("Unexpected AEF supplemental input dimensions")
    if int(alignment["global_bh_q_lt_0.05"].sum()) != 224 or int(alignment["global_bonferroni_p_lt_0.05"].sum()) != 44:
        raise RuntimeError("AEF--GEE site-alignment counts changed")
    if corrected_null.groupby(["pfam", "latent_axis"]).size().nunique() != 1 or int(corrected_null.groupby(["pfam", "latent_axis"]).size().iloc[0]) != 3:
        raise RuntimeError("Corrected AEF null is not three representations per pair")
    if int(corrected_candidates["robust_candidate_all_required_checks"].sum()) != 2:
        raise RuntimeError("Corrected AEF all-seven count changed")
    robust_pairs = set(
        map(
            tuple,
            corrected_candidates.loc[
                corrected_candidates["robust_candidate_all_required_checks"], ["pfam", "latent_axis"]
            ].to_numpy(),
        )
    )
    if robust_pairs != {("PF01638", "A52"), ("PF01638", "A53")}:
        raise RuntimeError(f"Unexpected corrected AEF robust pairs: {robust_pairs}")
    return alignment, variable_summary, axis_alignment, corrected_null, corrected_candidates, {
        "alignment_tests": len(alignment),
        "alignment_global_bh_q_lt_0.05": 224,
        "alignment_global_bonferroni_p_lt_0.05": 44,
        "corrected_null_rows": len(corrected_null),
        "corrected_selected_pairs": len(corrected_candidates),
        "corrected_all_seven": 2,
        "corrected_all_seven_pairs": sorted([list(pair) for pair in robust_pairs]),
        "permutations": int(corrected_null["permutations"].unique().item()),
    }


def build_s3(
    site_metadata: pd.DataFrame,
    site_axis_summary: pd.DataFrame,
    alignment: pd.DataFrame,
    alignment_variables: pd.DataFrame,
    alignment_axes: pd.DataFrame,
    corrected_null: pd.DataFrame,
    corrected_candidates: pd.DataFrame,
    site_audit: dict[str, Any],
    site_counts: dict[str, Any],
    aef_counts: dict[str, Any],
) -> tuple[dict[str, pd.DataFrame], list[str]]:
    workbook = load_workbook(S3_SOURCE)
    retained = ["S3A_pooled_AEF_Pfam", "S3B_phylum_centered", "S3F_within_phylum_summary"]
    pooled_pairs = read_sheet(S3_SOURCE, "S3A_pooled_AEF_Pfam")
    centered_pairs = read_sheet(S3_SOURCE, "S3B_phylum_centered")
    pooled_run_manifest = json.loads(AEF_RUN_MANIFEST.read_text(encoding="utf-8"))
    pooled_summary = pooled_run_manifest["computed_summary"]
    pfam_profiles = int(pooled_summary["strict_variable_pfam_profiles"])
    aef_axes = int(pooled_summary["aef_axes"])
    pooled_tests = int(pooled_summary["global_tests"])
    robust_pair_text = " and ".join(
        f"{pfam}–{axis}" for pfam, axis in aef_counts["corrected_all_seven_pairs"]
    )
    old_index = read_sheet(S3_SOURCE, "S3E_AEF_output_index")
    for title in ("README", "S3C_AEF_recorded_metadata", "S3D_AEF_axis_summary", "S3E_AEF_output_index"):
        workbook.remove(workbook[title])

    readme = pd.DataFrame(
        [
            ("Supplemental Table S3", "Secondary AlphaEarth Foundations (AEF) analyses and site-level external alignment with named GEE variables."),
            ("Analysis cohort", f"{site_audit['exact_id_genomes']} exact-ID macroalgal genomes; site-level analyses give each of {site_audit['unique_coordinate_sites']} unique coordinate sites one observation."),
            ("Pooled family", f"{pfam_profiles:,} strict Pfam accessions × {aef_axes} unitless AEF axes = {pooled_tests:,} tests."),
            ("Pooled result", f"{len(pooled_pairs)} pooled Pfam–AEF pairs met global BH q < 0.05."),
            ("S3A", f"All {len(pooled_pairs)} pooled raw-count Pfam–AEF pairs meeting global BH q < 0.05."),
            ("S3B", f"All {len(centered_pairs)} phylum-centered Pfam–AEF pairs meeting global BH q < 0.05."),
            ("S3C", f"All {site_counts['tests']} unique-site AEF correlations with recorded temperature, latitude, and longitude; BH and Bonferroni correction span {site_counts['tests']} tests."),
            ("S3D", "Unique-site recorded-metadata summary for each AEF axis, joined to the pooled Pfam summary. S3C–D supersede the prior genome-weighted metadata summaries."),
            ("S3E", "File-level provenance, SHA-256 hashes, byte sizes, and row counts."),
            ("S3F", "Within-phylum pooled Pfam--AEF screen summaries under the stated nonzero-prevalence specifications."),
            ("S3G", f"All {aef_counts['alignment_tests']} unique-site correlations between {aef_axes} unitless AEF axes and {len(alignment_variables)} named GEE variables; {aef_counts['alignment_global_bh_q_lt_0.05']} met global BH q < 0.05 and {aef_counts['alignment_global_bonferroni_p_lt_0.05']} met Bonferroni p < 0.05."),
            ("S3H", f"Named-GEE-variable summary of the {aef_counts['alignment_tests']}-pair site-level AEF–GEE alignment."),
            ("S3I", f"AEF-axis summary of the {aef_counts['alignment_tests']}-pair site-level AEF–GEE alignment."),
            ("S3J", f"Corrected structured-null results for {aef_counts['corrected_selected_pairs']} selected AEF–Pfam pairs across three abundance representations using {aef_counts['permutations']:,} conditional permutations per representation."),
            ("S3K", f"Corrected seven-check candidate summary for {aef_counts['corrected_selected_pairs']} selected AEF–Pfam pairs; {robust_pair_text} passed all seven checks."),
            ("Site aggregation", f"AEF axes and coordinates were identical within each exact coordinate site. Recorded temperature was averaged across genomes at a site; {site_audit['sites_with_multiple_recorded_temperatures']} sites contained more than one recorded temperature."),
            ("Axis interpretation", "S3G–I provide a site-level external crosswalk. A00–A63 remain unitless latent axes and are not assigned unique physical labels."),
        ],
        columns=["Item", "Statement"],
    )

    index_rows: list[dict[str, Any]] = []
    for row in old_index.itertuples(index=False):
        path = AEF_RUN_DIR / row.file
        if not path.is_file() or sha256(path) != row.sha256 or path.stat().st_size != int(row.bytes):
            raise RuntimeError(f"Existing AEF output-index entry failed authentication: {path}")
        index_rows.append(file_record("Pooled AEF run output", path, "authenticated existing index entry"))
    index_rows.extend(
        [
            file_record("Pooled AEF run manifest", AEF_RUN_MANIFEST, "validated"),
            file_record("Reconciled exact-ID analysis manifest", RECONCILED_MANIFEST, f"{site_audit['exact_id_genomes']} selected genomes"),
            file_record("AEF embeddings", AEF_EMBEDDINGS, f"{site_audit['exact_id_genomes']} genomes x {aef_axes} axes"),
            file_record("Site-level AEF--GEE alignment", AEF_ALIGNMENT, f"{len(alignment)} rows"),
            file_record("AEF--GEE named-variable summary", AEF_ALIGNMENT_VARIABLES, f"{len(alignment_variables)} rows"),
            file_record("AEF--GEE axis summary", AEF_ALIGNMENT_AXES, f"{len(alignment_axes)} rows"),
            file_record("AEF--GEE alignment manifest", AEF_ALIGNMENT_MANIFEST, "validated"),
            file_record("Corrected selected-AEF structured null", AEF_CORRECTED_NULL, f"{len(corrected_null)} rows"),
            file_record("Corrected selected-AEF candidate summary", AEF_CORRECTED_CANDIDATES, f"{len(corrected_candidates)} rows"),
            file_record("Corrected selected-AEF manifest", AEF_CORRECTED_MANIFEST, "validated"),
            file_record("Corrected selected-AEF generator", AEF_CORRECTION_SCRIPT, "executed"),
            file_record("Workbook source version", S3_SOURCE, "preserved"),
            file_record("Workbook builder", SCRIPT, "executed"),
        ]
    )
    file_index = pd.DataFrame(index_rows).sort_values(["file_role", "relative_path"], kind="stable").reset_index(drop=True)

    add_frame_sheet(workbook, "README", readme, index=0)
    add_frame_sheet(workbook, "S3C_site_recorded_metadata", site_metadata, index=3)
    add_frame_sheet(workbook, "S3D_site_axis_summary", site_axis_summary, index=4)
    add_frame_sheet(workbook, "S3E_output_index", file_index, index=5)
    add_frame_sheet(workbook, "S3G_AEF_GEE_alignment", alignment)
    add_frame_sheet(workbook, "S3H_GEE_variable_summary", alignment_variables)
    add_frame_sheet(workbook, "S3I_AEF_axis_alignment", alignment_axes)
    add_frame_sheet(workbook, "S3J_corrected_AEF_null", corrected_null)
    add_frame_sheet(workbook, "S3K_AEF_candidate_checks", corrected_candidates)
    workbook.properties.title = "Supplemental Table S3: secondary AEF analyses and site-level alignment"
    workbook.properties.creator = "Reproducible revision pipeline"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    if S3_OUTPUT.exists():
        raise FileExistsError(S3_OUTPUT)
    workbook.save(S3_OUTPUT)
    expected_frames = {
        "README": readme,
        "S3C_site_recorded_metadata": site_metadata,
        "S3D_site_axis_summary": site_axis_summary,
        "S3E_output_index": file_index,
        "S3G_AEF_GEE_alignment": alignment,
        "S3H_GEE_variable_summary": alignment_variables,
        "S3I_AEF_axis_alignment": alignment_axes,
        "S3J_corrected_AEF_null": corrected_null,
        "S3K_AEF_candidate_checks": corrected_candidates,
    }
    return expected_frames, retained


def canonical_cell(cell) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, datetime):
        value = value.isoformat()
    return {
        "coordinate": cell.coordinate,
        "value": value,
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "font": str(cell.font),
        "fill": str(cell.fill),
        "border": str(cell.border),
        "alignment": str(cell.alignment),
        "protection": str(cell.protection),
    }


def sheet_digest(workbook_path: Path, title: str) -> str:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet = workbook[title]
    payload = [
        canonical_cell(cell)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)
        for cell in row
    ]
    workbook.close()
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def audit_frame(workbook_path: Path, title: str, expected: pd.DataFrame) -> dict[str, Any]:
    observed = pd.read_excel(workbook_path, sheet_name=title)
    if list(observed.columns) != list(expected.columns):
        raise AssertionError(f"Column mismatch in {title}")
    pd.testing.assert_frame_equal(
        observed,
        expected.reset_index(drop=True),
        check_dtype=False,
        check_exact=False,
        rtol=1e-12,
        atol=1e-14,
        check_like=False,
    )
    return {
        "rows": len(expected),
        "columns": len(expected.columns),
        "source_frame_sha256": hashlib.sha256(
            pd.util.hash_pandas_object(expected.reset_index(drop=True), index=True).values.tobytes()
        ).hexdigest(),
        "roundtrip_frame_sha256": hashlib.sha256(
            pd.util.hash_pandas_object(observed.reset_index(drop=True), index=True).values.tobytes()
        ).hexdigest(),
        "roundtrip_values_equal_with_tolerance": True,
    }


def audit_workbook(
    source: Path,
    output: Path,
    expected_frames: dict[str, pd.DataFrame],
    retained_sheets: list[str],
    expected_sheetnames: list[str],
) -> dict[str, Any]:
    with zipfile.ZipFile(output, "r") as archive:
        bad_member = archive.testzip()
    workbook = load_workbook(output, read_only=True, data_only=False)
    observed_sheetnames = workbook.sheetnames
    formula_cells = sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    sheet_rows = {sheet.title: sheet.max_row - 1 for sheet in workbook.worksheets}
    workbook.close()
    if bad_member is not None:
        raise RuntimeError(f"Corrupt ZIP member in {output}: {bad_member}")
    if observed_sheetnames != expected_sheetnames:
        raise RuntimeError(f"Unexpected sheet order for {output}: {observed_sheetnames}")
    retained = {}
    for title in retained_sheets:
        source_digest = sheet_digest(source, title)
        output_digest = sheet_digest(output, title)
        retained[title] = {
            "source_cell_and_style_sha256": source_digest,
            "output_cell_and_style_sha256": output_digest,
            "unchanged": source_digest == output_digest,
        }
        if source_digest != output_digest:
            raise RuntimeError(f"Retained sheet changed: {title}")
    generated = {
        title: audit_frame(output, title, frame)
        for title, frame in expected_frames.items()
    }
    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "sheet_names": observed_sheetnames,
        "sheet_data_rows": sheet_rows,
        "formula_cells": formula_cells,
        "retained_sheet_audit": retained,
        "generated_sheet_roundtrip_audit": generated,
    }


def input_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "relative_path": relative(path),
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
    }


def write_integrity(
    path: Path,
    output: Path,
    source: Path,
    inputs: list[Path],
    analysis_counts: dict[str, Any],
    workbook_audit: dict[str, Any],
) -> None:
    if path.exists():
        raise FileExistsError(path)
    record = {
        "schema": "supplemental_workbook_scientific_integrity_v1",
        "status": "PASS",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "generator": input_record(SCRIPT),
        "software": {
            "python": platform.python_version(),
            "numpy": np.__version__,
            "pandas": pd.__version__,
            "scipy": scipy.__version__,
            "statsmodels": statsmodels.__version__,
            "openpyxl": openpyxl.__version__,
        },
        "data_integrity_policy": {
            "real_data_only": True,
            "synthetic_scientific_values": False,
            "hardcoded_scientific_results": False,
            "randomness_in_workbook_builder": False,
            "numerical_values_trace_to_inputs_or_declared_computation": True,
        },
        "source_workbook": input_record(source),
        "inputs": {item.name: input_record(item) for item in inputs},
        "analysis_counts": analysis_counts,
        "workbook_audit": workbook_audit,
        "output": input_record(output),
    }
    path.write_text(json.dumps(record, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")


def main() -> int:
    authenticate_fixed_inputs()
    for path in (S2_OUTPUT, S3_OUTPUT, S2_INTEGRITY, S3_INTEGRITY):
        if path.exists():
            raise FileExistsError(f"Refusing to overwrite {path}")

    sensitivity, gee_null, gee_candidates, gee_counts = validate_gee_inputs()
    s2_frames, s2_retained = build_s2(sensitivity, gee_null, gee_candidates, gee_counts)
    s2_sheetnames = [
        "README",
        "Table S2A Exact-ID GEE",
        "Table S2B Summary",
        "Table S2C FDR pairs",
        "Table S2D File Index",
        "S2E GEE sensitivity",
        "S2F Structured null",
        "S2G Candidate summary",
    ]
    s2_audit = audit_workbook(S2_SOURCE, S2_OUTPUT, s2_frames, s2_retained, s2_sheetnames)
    write_integrity(
        S2_INTEGRITY,
        S2_OUTPUT,
        S2_SOURCE,
        [
            GEE_DISCOVERY_FULL,
            GEE_EXTRACTION_MANIFEST,
            GEE_DISCOVERY_MANIFEST,
            GEE_SENSITIVITY,
            GEE_BASE_MANIFEST,
            GEE_REFINED_NULL,
            GEE_REFINED_CANDIDATES,
            GEE_REFINED_MANIFEST,
        ],
        gee_counts,
        s2_audit,
    )

    site, site_audit = load_site_metadata()
    site_metadata, site_axis_summary, site_counts = site_metadata_screen(site)
    alignment, alignment_variables, alignment_axes, corrected_null, corrected_candidates, aef_counts = validate_aef_inputs()
    s3_frames, s3_retained = build_s3(
        site_metadata,
        site_axis_summary,
        alignment,
        alignment_variables,
        alignment_axes,
        corrected_null,
        corrected_candidates,
        site_audit,
        site_counts,
        aef_counts,
    )
    s3_sheetnames = [
        "README",
        "S3A_pooled_AEF_Pfam",
        "S3B_phylum_centered",
        "S3C_site_recorded_metadata",
        "S3D_site_axis_summary",
        "S3E_output_index",
        "S3F_within_phylum_summary",
        "S3G_AEF_GEE_alignment",
        "S3H_GEE_variable_summary",
        "S3I_AEF_axis_alignment",
        "S3J_corrected_AEF_null",
        "S3K_AEF_candidate_checks",
    ]
    s3_audit = audit_workbook(S3_SOURCE, S3_OUTPUT, s3_frames, s3_retained, s3_sheetnames)
    write_integrity(
        S3_INTEGRITY,
        S3_OUTPUT,
        S3_SOURCE,
        [
            AEF_RUN_MANIFEST,
            RECONCILED_MANIFEST,
            AEF_EMBEDDINGS,
            AEF_ALIGNMENT,
            AEF_ALIGNMENT_VARIABLES,
            AEF_ALIGNMENT_AXES,
            AEF_ALIGNMENT_MANIFEST,
            AEF_CORRECTED_NULL,
            AEF_CORRECTED_CANDIDATES,
            AEF_CORRECTED_MANIFEST,
            AEF_CORRECTION_SCRIPT,
        ],
        {"site_audit": site_audit, "site_metadata": site_counts, "aef_outputs": aef_counts},
        s3_audit,
    )

    print(
        json.dumps(
            {
                "table_s2": {
                    "workbook": str(S2_OUTPUT.resolve()),
                    "integrity": str(S2_INTEGRITY.resolve()),
                    "sha256": sha256(S2_OUTPUT),
                    "analysis_counts": gee_counts,
                },
                "table_s3": {
                    "workbook": str(S3_OUTPUT.resolve()),
                    "integrity": str(S3_INTEGRITY.resolve()),
                    "sha256": sha256(S3_OUTPUT),
                    "site_audit": site_audit,
                    "site_metadata_counts": site_counts,
                    "aef_counts": aef_counts,
                },
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
