#!/usr/bin/env python3
"""Create Table S2 V6 by expanding the README's first SHA-256 use.

The source V5 workbook remains unchanged. Exactly one cell value may change;
all other values, all numeric cells, formulas, styles, and sheet structure are
required to remain identical.

Created: 2026-07-19 21:50:20 Asia/Bangkok
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook


SCRIPT_VERSION = "2026-07-19.1"
STAMP = "20260719_215020"
SCRIPT = Path(__file__).resolve()
SUPPLEMENT = SCRIPT.parent
ROOT = SCRIPT.parents[2]

SOURCE = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_214300_V5.xlsx"
SOURCE_INTEGRITY = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_214300_V5_integrity.json"
OUTPUT = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V6.xlsx"
INTEGRITY = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V6_integrity.json"

EXPECTED_SOURCE_SHA256 = "4e4cad63245d2dc56ac0afe32d111e679da3199025feac7e75f939e42126d88c"
EXPECTED_SOURCE_INTEGRITY_SHA256 = "a44338d02a5e04d1c67f29d16c18f22998486ccf2f1fe50ddc02e6da62935e91"
SOURCE_TEXT = "File-level provenance, SHA-256 hashes, byte sizes, and row counts."
OUTPUT_TEXT = (
    "File-level provenance, 256-bit Secure Hash Algorithm (SHA-256) digests, "
    "byte sizes, and row counts."
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT))


def authenticate() -> None:
    expected = {
        SOURCE: EXPECTED_SOURCE_SHA256,
        SOURCE_INTEGRITY: EXPECTED_SOURCE_INTEGRITY_SHA256,
    }
    for path, digest in expected.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        observed = sha256(path)
        if observed != digest:
            raise RuntimeError(f"Authenticated source changed: {path}: {observed}")
    source_integrity = json.loads(SOURCE_INTEGRITY.read_text(encoding="utf-8"))
    if source_integrity.get("status") != "PASS" or source_integrity.get("audit", {}).get("result") != "PASS":
        raise RuntimeError("Source V5 integrity record is not PASS")
    for path in (OUTPUT, INTEGRITY):
        if path.exists():
            raise FileExistsError(f"Refusing to overwrite {path}")


def create_output() -> None:
    workbook = load_workbook(SOURCE)
    cell = workbook["README"]["B9"]
    if cell.value != SOURCE_TEXT:
        raise RuntimeError(f"Unexpected README!B9 source: {cell.value!r}")
    cell.value = OUTPUT_TEXT
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(OUTPUT)


def value_equal(left: Any, right: Any) -> bool:
    if isinstance(left, float) and math.isnan(left):
        return isinstance(right, float) and math.isnan(right)
    return left == right


def style_payload(cell) -> tuple[Any, ...]:
    return (
        cell.number_format,
        str(cell.font),
        str(cell.fill),
        str(cell.border),
        str(cell.alignment),
        str(cell.protection),
    )


def audit_output() -> dict[str, Any]:
    with zipfile.ZipFile(OUTPUT, "r") as archive:
        bad_member = archive.testzip()
    if bad_member is not None:
        raise RuntimeError(f"ZIP CRC failure: {bad_member}")

    source_workbook = load_workbook(SOURCE, read_only=False, data_only=False)
    output_workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    if source_workbook.sheetnames != output_workbook.sheetnames:
        raise RuntimeError("Sheet names/order changed")

    value_changes = []
    style_changes = []
    formula_changes = []
    numeric_cells = 0
    cells_compared = 0
    for title in source_workbook.sheetnames:
        source_sheet = source_workbook[title]
        output_sheet = output_workbook[title]
        if (source_sheet.max_row, source_sheet.max_column) != (
            output_sheet.max_row,
            output_sheet.max_column,
        ):
            raise RuntimeError(f"Sheet dimensions changed: {title}")
        for row in source_sheet.iter_rows():
            for source_cell in row:
                cells_compared += 1
                output_cell = output_sheet[source_cell.coordinate]
                if isinstance(source_cell.value, (int, float)) and not isinstance(source_cell.value, bool):
                    numeric_cells += 1
                    if not value_equal(source_cell.value, output_cell.value):
                        raise RuntimeError(
                            f"Numeric cell changed: {title}!{source_cell.coordinate}"
                        )
                if not value_equal(source_cell.value, output_cell.value):
                    value_changes.append(
                        {
                            "sheet": title,
                            "cell": source_cell.coordinate,
                            "source": source_cell.value,
                            "output": output_cell.value,
                        }
                    )
                if style_payload(source_cell) != style_payload(output_cell):
                    style_changes.append(f"{title}!{source_cell.coordinate}")
                if source_cell.data_type == "f" or output_cell.data_type == "f":
                    if (source_cell.data_type, source_cell.value) != (
                        output_cell.data_type,
                        output_cell.value,
                    ):
                        formula_changes.append(f"{title}!{source_cell.coordinate}")

    source_workbook.close()
    output_workbook.close()
    expected_change = {
        "sheet": "README",
        "cell": "B9",
        "source": SOURCE_TEXT,
        "output": OUTPUT_TEXT,
    }
    if value_changes != [expected_change]:
        raise RuntimeError(f"Unexpected value changes: {value_changes}")
    if style_changes or formula_changes:
        raise RuntimeError(
            f"Style/formula changes found: styles={style_changes[:10]}, formulas={formula_changes[:10]}"
        )

    check = load_workbook(OUTPUT, read_only=True, data_only=False)
    sha_cells = []
    for sheet in check.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "SHA-256" in cell.value:
                    sha_cells.append((sheet.title, cell.coordinate, cell.value))
    check.close()
    if sha_cells != [("README", "B9", OUTPUT_TEXT)]:
        raise RuntimeError(f"SHA-256 expansion audit failed: {sha_cells}")

    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "sheet_names_and_dimensions_unchanged": True,
        "cells_compared": cells_compared,
        "numeric_cells_compared": numeric_cells,
        "numeric_differences": 0,
        "value_changes": value_changes,
        "style_changes": style_changes,
        "formula_changes": formula_changes,
        "sha256_abbreviation_occurrences": len(sha_cells),
        "sha256_first_use_expanded": True,
    }


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "relative_path": relative(path),
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
    }


def write_integrity(audit: dict[str, Any]) -> None:
    if INTEGRITY.exists():
        raise FileExistsError(INTEGRITY)
    record = {
        "schema": "supplemental_workbook_single_cell_integrity_v1",
        "status": "PASS",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "generator": file_record(SCRIPT),
        "software": {
            "python": platform.python_version(),
            "openpyxl": openpyxl.__version__,
        },
        "source_workbook": file_record(SOURCE),
        "source_integrity": file_record(SOURCE_INTEGRITY),
        "audit": audit,
        "output": file_record(OUTPUT),
    }
    INTEGRITY.write_text(
        json.dumps(record, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def main() -> int:
    authenticate()
    create_output()
    audit = audit_output()
    write_integrity(audit)
    print(
        json.dumps(
            {
                "workbook": str(OUTPUT.resolve()),
                "sha256": sha256(OUTPUT),
                "integrity": str(INTEGRITY.resolve()),
                "only_changed_cell": "README!B9",
                "numeric_cells_unchanged": audit["numeric_cells_compared"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
