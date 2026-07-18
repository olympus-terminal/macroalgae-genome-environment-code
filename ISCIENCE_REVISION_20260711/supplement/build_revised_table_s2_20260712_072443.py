#!/usr/bin/env python3
"""Build revised Table S2 from the corrected exact-ID GEE validation.

The original workbook is preserved unchanged. This version excludes results
from the non-unique species-name merge and includes only the authenticated
126-ID extraction, corrected raw-count summary, FDR-significant pairs, and a
hash index to the complete compressed result table.
"""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


ROOT = Path(__file__).resolve().parents[2]
OUTDIR = ROOT / "ISCIENCE_REVISION_20260711" / "supplement"
ORIGINAL = ROOT / "TABLES" / "Table_S2_GEE.xlsx"
EXTRACTION = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_environmental_extraction_20260712_071838.csv"
EXTRACTION_MANIFEST = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_environmental_extraction_manifest_20260712_071838.json"
ALL_RESULTS = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_raw_pfam_correlations_20260712_072151.csv.gz"
SIGNIFICANT = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_raw_pfam_fdr05_20260712_072151.csv"
SUMMARY = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_raw_pfam_summary_20260712_072151.csv"
VALIDATION_MANIFEST = ROOT / "ISCIENCE_REVISION_20260711/gee_validation/exact_id_gee_correlation_validation_manifest_20260712_072151.json"
OUTPUT = OUTDIR / "Table_S2_revised_exact_id_GEE_validation_20260712_072443.xlsx"
MANIFEST = OUTDIR / "Table_S2_revised_exact_id_GEE_validation_integrity_20260712_072443.json"

HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEAD_FILL = PatternFill("solid", fgColor="D9EAD3")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def style_sheet(ws, *, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cells in ws.iter_cols():
        letter = get_column_letter(cells[0].column)
        sample = cells[: min(len(cells), 200)]
        width = max((len(str(cell.value)) for cell in sample if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")


def add_dataframe(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title)
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)
    style_sheet(ws)


def main() -> None:
    inputs = [
        ORIGINAL,
        EXTRACTION,
        EXTRACTION_MANIFEST,
        ALL_RESULTS,
        SIGNIFICANT,
        SUMMARY,
        VALIDATION_MANIFEST,
    ]
    for path in inputs:
        if not path.is_file():
            raise FileNotFoundError(path)

    extraction_manifest = json.loads(EXTRACTION_MANIFEST.read_text(encoding="utf-8"))
    validation_manifest = json.loads(VALIDATION_MANIFEST.read_text(encoding="utf-8"))
    if sha256(EXTRACTION) != extraction_manifest["output"]["sha256"]:
        raise RuntimeError("Extraction does not match its manifest")
    for relative, expected in validation_manifest["outputs"].items():
        path = ROOT / relative
        if sha256(path) != expected:
            raise RuntimeError(f"Validation output hash mismatch: {path}")

    extraction = pd.read_csv(EXTRACTION, low_memory=False)
    significant = pd.read_csv(SIGNIFICANT, low_memory=False)
    summary = pd.read_csv(SUMMARY, low_memory=False)
    if len(extraction) != 126 or extraction["genome_id"].nunique() != 126:
        raise RuntimeError("Extraction is not 126 unique genome IDs")
    if len(significant) != validation_manifest["results"]["fdr_lt_0_05"]:
        raise RuntimeError("Significant-pair count differs from validation manifest")
    if summary.iloc[-1]["variable"] != "TOTAL":
        raise RuntimeError("Summary TOTAL row absent")
    total_tests = int(validation_manifest["results"]["total_tests"])
    significant_pairs = len(significant)

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    notes = [
        ("Table S2 revision", "Exact-genome-ID GEE validation of the original discovery workflow."),
        ("Correction made", "The submitted workflow joined environmental and Pfam tables using non-unique species names, producing a many-to-many expansion. This revised table uses all 126 canonical genome IDs as the only join key."),
        ("Scope", "This is a bounded data-integrity correction of the original analysis, not a new discovery screen."),
        ("Primary abundance", "Reconstructed raw HMM Pfam counts. No BUSCO or denominator rescaling is applied in this corrected discovery table; those specifications are reported separately as post hoc sensitivities."),
        ("Prevalence rule", "Pfams present in at least 7 of 126 genomes (the prespecified 5% rule) were retained before environmental testing."),
        ("Multiplicity", f"Benjamini–Hochberg and Bonferroni correction were applied across the complete {total_tests:,}-pair raw-count GEE family."),
        ("Table S2A", "Authenticated extraction of 13 environmental variables for 126 exact genome IDs."),
        ("Table S2B", "Corrected result counts by environmental variable and for the full test family."),
        ("Table S2C", f"All {significant_pairs:,} pairs with global Benjamini–Hochberg q < 0.05."),
        ("Table S2D", f"Hash index to the complete {total_tests:,}-row compressed result table and provenance manifests."),
        ("Original workbook preserved", str(ORIGINAL.relative_to(ROOT))),
        ("Original workbook SHA-256", sha256(ORIGINAL)),
    ]
    readme.append(["Item", "Statement"])
    for item, statement in notes:
        readme.append([item, statement])
    style_sheet(readme)
    readme.column_dimensions["A"].width = 30
    readme.column_dimensions["B"].width = 120
    for row in readme.iter_rows(min_row=2):
        row[0].fill = SUBHEAD_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    add_dataframe(workbook, "Table S2A Exact-ID GEE", extraction)
    add_dataframe(workbook, "Table S2B Summary", summary)
    add_dataframe(workbook, "Table S2C FDR pairs", significant)

    index = workbook.create_sheet("Table S2D File Index")
    index.append(["File role", "Relative path", "SHA-256", "Rows/status"])
    file_rows = [
        ("Complete corrected correlation family", ALL_RESULTS, f"{total_tests:,} rows"),
        ("Exact-ID extraction provenance", EXTRACTION_MANIFEST, "validated"),
        ("Corrected correlation provenance", VALIDATION_MANIFEST, "validated"),
    ]
    for role, path, status in file_rows:
        index.append([role, str(path.relative_to(ROOT)), sha256(path), status])
    style_sheet(index)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)

    check = load_workbook(OUTPUT, read_only=True, data_only=True)
    expected_sheets = [
        "README",
        "Table S2A Exact-ID GEE",
        "Table S2B Summary",
        "Table S2C FDR pairs",
        "Table S2D File Index",
    ]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected sheet sequence: {check.sheetnames}")
    if check[expected_sheets[1]].max_row != len(extraction) + 1:
        raise RuntimeError("Materialized S2A row count incorrect")
    if check[expected_sheets[3]].max_row != significant_pairs + 1:
        raise RuntimeError("Materialized S2C row count incorrect")
    check.close()

    manifest = {
        "generated_by": str(Path(__file__).relative_to(ROOT)),
        "generator_sha256": sha256(Path(__file__)),
        "python_version": sys.version,
        "openpyxl_version": openpyxl.__version__,
        "output": str(OUTPUT.relative_to(ROOT)),
        "output_sha256": sha256(OUTPUT),
        "inputs": {str(path.relative_to(ROOT)): sha256(path) for path in inputs},
        "checks": {
            "exact_126_id_extraction": True,
            "corrected_total_tests": validation_manifest["results"]["total_tests"],
            "corrected_fdr_pairs": validation_manifest["results"]["fdr_lt_0_05"],
            "corrected_bonferroni_pairs": validation_manifest["results"]["bonferroni_lt_0_05"],
            "old_many_to_many_results_excluded": True,
            "no_imputation": True,
            "no_synthetic_values": True,
        },
    }
    MANIFEST.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(OUTPUT), "manifest": str(MANIFEST), **manifest["checks"]}, indent=2))


if __name__ == "__main__":
    main()
