#!/usr/bin/env python3
"""Build Table S2 V4 and Table S3 V7 without changing prior versions.

Table S2 V4 appends authenticated, accession-level InterPro annotation fields
to the 84-row GEE candidate sheet and includes the 30-record annotation source
table. Only the 49 all-seven retained pairs receive annotations; those pairs
represent exactly 30 distinct Pfam accessions.

Table S3 V7 changes interpretation labels to "site-level cross-representation
alignment/crosswalk", explicitly describes the analysis as descriptive rather
than a decoder or independent validation, and clarifies the 126-ID to 90-site
aggregation. All pre-existing numeric cells are required to remain identical.

No synthetic or placeholder scientific data are used. Existing workbooks are
preserved and every output is audited after materialization.

Created: 2026-07-19 21:34:08 Asia/Bangkok
"""

from __future__ import annotations

import hashlib
import json
import math
import platform
import sys
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_VERSION = "2026-07-19.1"
STAMP = "20260719_213408"
SCRIPT = Path(__file__).resolve()
SUPPLEMENT = SCRIPT.parent
ROOT = SCRIPT.parents[2]

S2_SOURCE = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_211551_V3.xlsx"
S2_SOURCE_INTEGRITY = SUPPLEMENT / "Table_S2_revised_exact_id_GEE_validation_20260719_211551_V3_integrity.json"
S3_SOURCE = SUPPLEMENT / "Table_S3_AEF_20260719_211551_V6.xlsx"
S3_SOURCE_INTEGRITY = SUPPLEMENT / "Table_S3_AEF_20260719_211551_V6_integrity.json"

ANNOTATION_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "annotations"
ANNOTATION_CSV = ANNOTATION_DIR / "GEE_all_seven_InterPro_annotations_20260719_212508.csv"
ANNOTATION_MANIFEST = ANNOTATION_DIR / "GEE_all_seven_InterPro_annotations_20260719_212508.json"
ANNOTATION_GENERATOR = ANNOTATION_DIR / "fetch_gee_robust_interpro_20260719_212508.py"
AEF_ALIGNMENT_MANIFEST = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "gee_validation"
    / "AEF_GEE_site_alignment_manifest_20260719_204925.json"
)

S2_OUTPUT = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V4.xlsx"
S2_INTEGRITY = SUPPLEMENT / f"Table_S2_revised_exact_id_GEE_validation_{STAMP}_V4_integrity.json"
S3_OUTPUT = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V7.xlsx"
S3_INTEGRITY = SUPPLEMENT / f"Table_S3_AEF_{STAMP}_V7_integrity.json"

EXPECTED_HASHES = {
    S2_SOURCE: "085c4bdb56cca9d396c8716dca4326778bdba5202dadb1ff21cd5dd6414d6e4e",
    S2_SOURCE_INTEGRITY: "8e44c4cab8adb407c32d358306250821f3048e7c6ee0eec6e348f493020614f5",
    S3_SOURCE: "db9b5c1cad208a6b624586e07ad0d85e87253e48f2e0fdeab367c9377a6cd69e",
    S3_SOURCE_INTEGRITY: "9d07149fc191e8c755079f04d986ca686228c854707568168933f18a00be5036",
    ANNOTATION_CSV: "78ed20376da91404aa8b06395f00dc86f65a84f75ba55d5af6436aa949fd5106",
    ANNOTATION_MANIFEST: "bbb90d399b06ab72946e243fe42f70d27b24bdb6bcee232146103cb43ee25108",
    ANNOTATION_GENERATOR: "b52c8d5f96feaf0e201678e3c5e10d30001126a049f43b4becf83175ac7cb8d3",
    AEF_ALIGNMENT_MANIFEST: "1574fb484dff0719a7b9c4d025dfef3b0c4968b58470be03be4c8f5631aab0a1",
}

S2_ORIGINAL_CANDIDATE_COLUMNS = [
    "pfam",
    "environment",
    "discovery_spearman_r",
    "discovery_global_bh_q",
    "discovery_bonferroni_p",
    "discovery_bonferroni_significant",
    "reference_raw_site_effect",
    "direction_consistent_required_checks",
    "selected_set_q_lt_0.05_raw_site",
    "selected_set_q_lt_0.05_total_site",
    "selected_set_q_lt_0.05_peptide_site",
    "selected_set_q_lt_0.05_total_quality",
    "selected_set_q_lt_0.05_total_busco50",
    "selected_set_q_lt_0.05_total_tree",
    "selected_set_q_lt_0.05_total_structured_null",
    "checks_passed",
    "checks_required",
    "robust_candidate_all_required_checks",
    "interpretation_boundary",
]

ANNOTATION_COLUMN_MAP = [
    ("interpro_short_name", "short_name"),
    ("interpro_name", "name"),
    ("interpro_entry_type", "entry_type"),
    ("integrated_interpro_accession", "integrated_interpro"),
    ("interpro_source_database", "source_database"),
    ("interpro_api_url", "api_url"),
    ("interpro_http_status", "http_status"),
    ("interpro_response_sha256", "response_sha256"),
]

S3_ALIGNMENT_BOUNDARY = (
    "descriptive site-level cross-representation alignment; it is not a decoder "
    "or independent validation and does not assign a unique physical label to an AEF axis"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT))


def authenticate_inputs() -> None:
    for path, expected in EXPECTED_HASHES.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        observed = sha256(path)
        if observed != expected:
            raise RuntimeError(f"Authenticated input changed: {path}: {observed}")
    for path in (S2_OUTPUT, S2_INTEGRITY, S3_OUTPUT, S3_INTEGRITY):
        if path.exists():
            raise FileExistsError(f"Refusing to overwrite {path}")
    for path in (S2_SOURCE_INTEGRITY, S3_SOURCE_INTEGRITY):
        record = json.loads(path.read_text(encoding="utf-8"))
        if record.get("status") != "PASS" or record.get("workbook_audit", {}).get("result") != "PASS":
            raise RuntimeError(f"Source integrity did not pass: {path}")


def sheet_frame(path: Path, title: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=title)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def style_frame_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 32
    headers = {cell.column: str(cell.value) for cell in sheet[1]}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            header = headers.get(cell.column, "").lower()
            wrap = any(token in header for token in ("name", "url", "statement", "path", "boundary"))
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000E+00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, cells in enumerate(sheet.iter_cols(), start=1):
        header = headers.get(index, "").lower()
        values = [str(cell.value) if cell.value is not None else "" for cell in cells[:500]]
        cap = 55 if any(token in header for token in ("name", "url", "statement", "path", "boundary")) else 38
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(max(map(len, values), default=0) + 2, 10), cap
        )


def add_frame_sheet(workbook, title: str, frame: pd.DataFrame, index: int | None = None) -> None:
    if title in workbook.sheetnames:
        raise RuntimeError(f"Sheet already exists: {title}")
    sheet = workbook.create_sheet(title, index) if index is not None else workbook.create_sheet(title)
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append([None if pd.isna(value) else value for value in row])
    style_frame_sheet(sheet)


def load_annotations(candidate_frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    annotation = pd.read_csv(ANNOTATION_CSV, low_memory=False)
    manifest = json.loads(ANNOTATION_MANIFEST.read_text(encoding="utf-8"))
    required = [
        "accession",
        "short_name",
        "name",
        "entry_type",
        "integrated_interpro",
        "source_database",
        "api_url",
        "http_status",
        "response_sha256",
    ]
    if list(annotation.columns) != required:
        raise RuntimeError(f"Unexpected annotation schema: {annotation.columns.tolist()}")
    if len(annotation) != 30 or annotation["accession"].nunique() != 30:
        raise RuntimeError("InterPro input is not 30 unique Pfam accessions")
    if not annotation["http_status"].eq(200).all() or not annotation["source_database"].eq("pfam").all():
        raise RuntimeError("InterPro annotation retrieval status/source changed")
    manifest_records = pd.DataFrame(manifest["records"])[required]
    pd.testing.assert_frame_equal(
        annotation.reset_index(drop=True), manifest_records.reset_index(drop=True), check_dtype=False
    )
    if manifest["program_sha256"] != sha256(ANNOTATION_GENERATOR):
        raise RuntimeError("Annotation generator hash differs from manifest")
    if manifest["candidate_input_sha256"] != "85bd71ff3128219949e74b168a8714f9b65f2020b5854a0ee4ff0f05059d2e76":
        raise RuntimeError("Annotation manifest identifies an unexpected candidate input")
    robust = candidate_frame.loc[candidate_frame["robust_candidate_all_required_checks"].eq(True)]
    robust_pfams = set(robust["pfam"])
    annotation_pfams = set(annotation["accession"])
    if len(robust) != 49 or len(robust_pfams) != 30 or robust_pfams != annotation_pfams:
        raise RuntimeError("The 49 retained pairs and 30 annotated Pfams do not match exactly")
    if manifest["robust_pair_count"] != len(robust) or manifest["distinct_pfam_count"] != len(robust_pfams):
        raise RuntimeError("Annotation manifest counts differ from candidate table")
    return annotation, {
        "candidate_rows": len(candidate_frame),
        "primary_retained_pairs": len(robust),
        "primary_retained_pfams": len(robust_pfams),
        "annotation_records": len(annotation),
        "http_200_records": int(annotation["http_status"].eq(200).sum()),
    }


def build_s2() -> dict[str, Any]:
    candidate = sheet_frame(S2_SOURCE, "S2G Candidate summary")
    if list(candidate.columns) != S2_ORIGINAL_CANDIDATE_COLUMNS or len(candidate) != 84:
        raise RuntimeError("Unexpected source candidate table")
    annotation, counts = load_annotations(candidate)
    annotation_lookup = annotation.set_index("accession")

    workbook = load_workbook(S2_SOURCE)
    sheet = workbook["S2G Candidate summary"]
    original_max_column = sheet.max_column
    if original_max_column != len(S2_ORIGINAL_CANDIDATE_COLUMNS):
        raise RuntimeError("Unexpected source candidate worksheet width")
    header_template = sheet.cell(1, original_max_column)
    data_template_column = original_max_column
    for offset, (output_name, _) in enumerate(ANNOTATION_COLUMN_MAP, start=1):
        cell = sheet.cell(1, original_max_column + offset, output_name)
        copy_cell_style(header_template, cell)
    robust_rows = 0
    annotated_pfams: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        pfam = sheet.cell(row_index, 1).value
        robust = sheet.cell(row_index, 18).value is True
        record = annotation_lookup.loc[pfam] if robust else None
        if robust:
            robust_rows += 1
            annotated_pfams.add(pfam)
        for offset, (_, source_name) in enumerate(ANNOTATION_COLUMN_MAP, start=1):
            cell = sheet.cell(row_index, original_max_column + offset)
            copy_cell_style(sheet.cell(row_index, data_template_column), cell)
            cell.value = record[source_name] if record is not None else None
            if source_name == "http_status" and cell.value is not None:
                cell.number_format = "0"
    if robust_rows != 49 or len(annotated_pfams) != 30:
        raise RuntimeError("Candidate-sheet annotation coverage changed")
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "interpro_short_name": 20,
        "interpro_name": 55,
        "interpro_entry_type": 18,
        "integrated_interpro_accession": 26,
        "interpro_source_database": 24,
        "interpro_api_url": 55,
        "interpro_http_status": 20,
        "interpro_response_sha256": 55,
    }
    for offset, (output_name, _) in enumerate(ANNOTATION_COLUMN_MAP, start=1):
        sheet.column_dimensions[get_column_letter(original_max_column + offset)].width = widths[output_name]
        if any(token in output_name for token in ("name", "url", "sha256")):
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row_index, original_max_column + offset).alignment = Alignment(vertical="top", wrap_text=True)

    old_readme = sheet_frame(S2_SOURCE, "README")
    old_readme.loc[old_readme["Item"].eq("S2G"), "Statement"] = (
        "All 84 candidates with verified accession-level InterPro fields appended for the 49 primary retained pairs."
    )
    insertion = pd.DataFrame(
        [
            {
                "Item": "Primary retained GEE result set",
                "Statement": (
                    f"The primary retained GEE result set comprises {counts['primary_retained_pairs']} "
                    f"Pfam–environment pairs across {counts['primary_retained_pfams']} Pfam accessions that "
                    "retained direction and selected-family support in all seven required specifications."
                ),
            }
        ]
    )
    discovery_position = old_readme.index[old_readme["Item"].eq("Discovery results")].item() + 1
    readme = pd.concat(
        [old_readme.iloc[:discovery_position], insertion, old_readme.iloc[discovery_position:]],
        ignore_index=True,
    )
    readme = pd.concat(
        [
            readme,
            pd.DataFrame(
                [
                    {
                        "Item": "S2H",
                        "Statement": (
                            f"One authenticated InterPro record for each of the {counts['primary_retained_pfams']} "
                            "Pfam accessions represented by the primary retained result set."
                        ),
                    },
                    {
                        "Item": "Annotation scope",
                        "Statement": (
                            "InterPro names and entry types describe Pfam accessions; they are accession-level "
                            "database labels rather than sequence-specific functional assignments."
                        ),
                    },
                ]
            ),
        ],
        ignore_index=True,
    )
    workbook.remove(workbook["README"])
    add_frame_sheet(workbook, "README", readme, index=0)

    file_index = workbook["Table S2D File Index"]
    existing_paths = {
        file_index.cell(row, 2).value for row in range(2, file_index.max_row + 1)
    }
    provenance = [
        ("Primary retained GEE InterPro annotations", ANNOTATION_CSV, "30 accession records"),
        ("InterPro annotation retrieval manifest", ANNOTATION_MANIFEST, "authenticated"),
        ("InterPro annotation retrieval program", ANNOTATION_GENERATOR, "executed"),
        ("Workbook source version V3", S2_SOURCE, "preserved"),
        ("Workbook source V3 integrity", S2_SOURCE_INTEGRITY, "PASS"),
        ("Workbook V4 builder", SCRIPT, "executed"),
    ]
    for role, path, status in provenance:
        rel = relative(path)
        if rel in existing_paths:
            continue
        row_index = file_index.max_row + 1
        values = [role, rel, sha256(path), path.stat().st_size, status]
        for column_index, value in enumerate(values, start=1):
            target = file_index.cell(row_index, column_index, value)
            copy_cell_style(file_index.cell(row_index - 1, column_index), target)
    file_index.auto_filter.ref = file_index.dimensions

    add_frame_sheet(workbook, "S2H Retained annotations", annotation)
    workbook.properties.title = "Supplemental Table S2: primary retained GEE results and verified annotations"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(S2_OUTPUT)
    return counts


def build_s3() -> dict[str, Any]:
    alignment_manifest = json.loads(AEF_ALIGNMENT_MANIFEST.read_text(encoding="utf-8"))
    site_audit = alignment_manifest["site_audit"]
    expected_audit = {
        "exact_id_genomes": 126,
        "unique_coordinate_sites": 90,
        "repeated_coordinate_sites": 19,
        "within_site_aef_disagreements": 0,
        "within_site_gee_disagreements": 0,
    }
    for key, expected in expected_audit.items():
        if site_audit.get(key) != expected:
            raise RuntimeError(f"AEF--GEE site audit changed for {key}: {site_audit.get(key)}")

    workbook = load_workbook(S3_SOURCE)
    readme = workbook["README"]
    readme_changes = {
        "B2": "Secondary AlphaEarth Foundations (AEF) analyses and descriptive site-level cross-representation alignment with named GEE variables.",
        "B3": "The 126 exact-ID genome records mapped to 90 unique coordinate sites; site-level analyses give each coordinate one observation.",
        "B12": "All 832 descriptive site-level cross-representation correlations between 64 unitless AEF axes and 13 named GEE variables; 224 met global BH q < 0.05 and 44 met Bonferroni p < 0.05.",
        "B13": "Named-GEE-variable summary of the 832-pair descriptive site-level cross-representation alignment.",
        "B14": "AEF-axis summary of the 832-pair descriptive site-level cross-representation alignment.",
        "B17": "For S3G–I, genome records sharing a coordinate had identical AEF values and identical GEE values, so one value per coordinate was retained. For S3C–D, recorded temperature was averaged across genome records at a site; 7 sites contained more than one recorded temperature.",
        "B18": "S3G–I report a descriptive site-level cross-representation alignment/crosswalk. It is not a decoder or independent validation; A00–A63 remain unitless and are not assigned unique physical labels.",
    }
    for coordinate, value in readme_changes.items():
        readme[coordinate] = value

    alignment = workbook["S3G_AEF_GEE_alignment"]
    if alignment.cell(1, 10).value != "interpretation_boundary" or alignment.max_row != 833:
        raise RuntimeError("Unexpected S3G alignment sheet")
    for row_index in range(2, alignment.max_row + 1):
        alignment.cell(row_index, 10).value = S3_ALIGNMENT_BOUNDARY
    alignment.column_dimensions["J"].width = 55
    for row_index in range(2, alignment.max_row + 1):
        alignment.cell(row_index, 10).alignment = Alignment(vertical="top", wrap_text=True)

    output_index = workbook["S3E_output_index"]
    role_replacements = {
        "AEF--GEE alignment manifest": "Site-level cross-representation alignment manifest",
        "AEF--GEE axis summary": "Site-level cross-representation axis summary",
        "AEF--GEE named-variable summary": "Site-level cross-representation named-variable summary",
        "Site-level AEF--GEE alignment": "Site-level cross-representation alignment",
    }
    observed_roles = set()
    for row_index in range(2, output_index.max_row + 1):
        cell = output_index.cell(row_index, 1)
        if cell.value in role_replacements:
            observed_roles.add(cell.value)
            cell.value = role_replacements[cell.value]
    if observed_roles != set(role_replacements):
        raise RuntimeError(f"Did not find all provenance role labels: {set(role_replacements) - observed_roles}")

    workbook.properties.title = "Supplemental Table S3: secondary AEF analyses and cross-representation alignment"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(S3_OUTPUT)
    return {
        **expected_audit,
        "alignment_rows_relabelled": alignment.max_row - 1,
        "readme_cells_updated": len(readme_changes),
        "provenance_role_labels_updated": len(role_replacements),
    }


def is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, float) and math.isnan(left):
        return isinstance(right, float) and math.isnan(right)
    return left == right


def digest_cells(path: Path, title: str, max_column: int | None = None) -> str:
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook[title]
    maximum_column = max_column or sheet.max_column
    payload = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=maximum_column):
        for cell in row:
            payload.append(
                (
                    cell.coordinate,
                    cell.value,
                    cell.data_type,
                    cell.number_format,
                    str(cell.font),
                    str(cell.fill),
                    str(cell.border),
                    str(cell.alignment),
                )
            )
    workbook.close()
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")).hexdigest()


def audit_existing_numeric_cells(source: Path, output: Path) -> dict[str, Any]:
    source_workbook = load_workbook(source, read_only=False, data_only=False)
    output_workbook = load_workbook(output, read_only=False, data_only=False)
    compared = 0
    differences = []
    for title in source_workbook.sheetnames:
        source_sheet = source_workbook[title]
        output_sheet = output_workbook[title]
        for row in source_sheet.iter_rows():
            for source_cell in row:
                if not is_numeric(source_cell.value):
                    continue
                compared += 1
                output_cell = output_sheet[source_cell.coordinate]
                if not values_equal(source_cell.value, output_cell.value):
                    differences.append(
                        {
                            "sheet": title,
                            "cell": source_cell.coordinate,
                            "source": source_cell.value,
                            "output": output_cell.value,
                        }
                    )
    source_workbook.close()
    output_workbook.close()
    if differences:
        raise RuntimeError(f"Existing numeric cells changed: {differences[:10]}")
    return {"numeric_cells_compared": compared, "differences": 0, "result": "PASS"}


def audit_s2(counts: dict[str, Any]) -> dict[str, Any]:
    with zipfile.ZipFile(S2_OUTPUT, "r") as archive:
        if archive.testzip() is not None:
            raise RuntimeError("S2 output ZIP failed CRC test")
    workbook = load_workbook(S2_OUTPUT, read_only=False, data_only=False)
    expected_sheets = [
        "README",
        "Table S2A Exact-ID GEE",
        "Table S2B Summary",
        "Table S2C FDR pairs",
        "Table S2D File Index",
        "S2E GEE sensitivity",
        "S2F Structured null",
        "S2G Candidate summary",
        "S2H Retained annotations",
    ]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected S2 V4 sheets: {workbook.sheetnames}")
    candidate = workbook["S2G Candidate summary"]
    headers = [candidate.cell(1, column).value for column in range(1, candidate.max_column + 1)]
    expected_headers = S2_ORIGINAL_CANDIDATE_COLUMNS + [item[0] for item in ANNOTATION_COLUMN_MAP]
    if headers != expected_headers or candidate.max_row != 85:
        raise RuntimeError("S2G output schema/dimensions changed")
    robust_annotated = 0
    nonrobust_nonblank = 0
    annotated_pfams: set[str] = set()
    for row_index in range(2, candidate.max_row + 1):
        robust = candidate.cell(row_index, 18).value is True
        values = [
            candidate.cell(row_index, len(S2_ORIGINAL_CANDIDATE_COLUMNS) + offset).value
            for offset in range(1, len(ANNOTATION_COLUMN_MAP) + 1)
        ]
        if robust:
            if any(value is None for value in values):
                raise RuntimeError(f"Missing annotation in robust S2G row {row_index}")
            robust_annotated += 1
            annotated_pfams.add(candidate.cell(row_index, 1).value)
        elif any(value is not None for value in values):
            nonrobust_nonblank += 1
    if (robust_annotated, len(annotated_pfams), nonrobust_nonblank) != (49, 30, 0):
        raise RuntimeError("S2G annotation coverage audit failed")
    workbook.close()

    annotation_roundtrip = sheet_frame(S2_OUTPUT, "S2H Retained annotations")
    annotation_source = pd.read_csv(ANNOTATION_CSV)
    pd.testing.assert_frame_equal(annotation_roundtrip, annotation_source, check_dtype=False)
    unchanged_sheets = [
        "Table S2A Exact-ID GEE",
        "Table S2B Summary",
        "Table S2C FDR pairs",
        "S2E GEE sensitivity",
        "S2F Structured null",
    ]
    retained_audit = {}
    for title in unchanged_sheets:
        source_digest = digest_cells(S2_SOURCE, title)
        output_digest = digest_cells(S2_OUTPUT, title)
        retained_audit[title] = {
            "source_sha256": source_digest,
            "output_sha256": output_digest,
            "unchanged": source_digest == output_digest,
        }
        if source_digest != output_digest:
            raise RuntimeError(f"Unchanged S2 sheet changed: {title}")
    candidate_source_digest = digest_cells(
        S2_SOURCE, "S2G Candidate summary", len(S2_ORIGINAL_CANDIDATE_COLUMNS)
    )
    candidate_output_digest = digest_cells(
        S2_OUTPUT, "S2G Candidate summary", len(S2_ORIGINAL_CANDIDATE_COLUMNS)
    )
    if candidate_source_digest != candidate_output_digest:
        raise RuntimeError("Original S2G cells/styles changed")
    numeric = audit_existing_numeric_cells(S2_SOURCE, S2_OUTPUT)
    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "sheet_names": expected_sheets,
        "annotation_counts": counts,
        "original_candidate_cells_unchanged": True,
        "original_candidate_payload_sha256": candidate_source_digest,
        "annotation_roundtrip_equal": True,
        "retained_sheet_audit": retained_audit,
        "existing_numeric_cell_audit": numeric,
    }


def audit_s3(change_counts: dict[str, Any]) -> dict[str, Any]:
    with zipfile.ZipFile(S3_OUTPUT, "r") as archive:
        if archive.testzip() is not None:
            raise RuntimeError("S3 output ZIP failed CRC test")
    source_workbook = load_workbook(S3_SOURCE, read_only=False, data_only=False)
    output_workbook = load_workbook(S3_OUTPUT, read_only=False, data_only=False)
    if output_workbook.sheetnames != source_workbook.sheetnames:
        raise RuntimeError("S3 sheet names/order changed")
    numeric = audit_existing_numeric_cells(S3_SOURCE, S3_OUTPUT)
    text_changes = []
    forbidden_external = []
    cross_representation_mentions = 0
    independent_validation_mentions = 0
    for title in source_workbook.sheetnames:
        source_sheet = source_workbook[title]
        output_sheet = output_workbook[title]
        for row in output_sheet.iter_rows():
            for output_cell in row:
                value = output_cell.value
                if isinstance(value, str):
                    lower = value.lower()
                    if "external alignment" in lower or "external crosswalk" in lower:
                        forbidden_external.append(f"{title}!{output_cell.coordinate}")
                    if "cross-representation" in lower:
                        cross_representation_mentions += 1
                    if "independent validation" in lower:
                        independent_validation_mentions += 1
        for row in source_sheet.iter_rows():
            for source_cell in row:
                output_value = output_sheet[source_cell.coordinate].value
                if not values_equal(source_cell.value, output_value):
                    text_changes.append(
                        {
                            "sheet": title,
                            "cell": source_cell.coordinate,
                            "source": source_cell.value,
                            "output": output_value,
                        }
                    )
    source_workbook.close()
    output_workbook.close()
    expected_text_changes = 7 + 832 + 4
    if len(text_changes) != expected_text_changes:
        raise RuntimeError(f"Unexpected number of S3 text changes: {len(text_changes)}")
    if forbidden_external:
        raise RuntimeError(f"External-alignment terminology remains: {forbidden_external[:10]}")
    if cross_representation_mentions < 832 + 7 or independent_validation_mentions < 833:
        raise RuntimeError("Required S3 interpretation wording is incomplete")
    return {
        "result": "PASS",
        "zip_crc_test": "PASS",
        "sheet_names_unchanged": True,
        "existing_numeric_cell_audit": numeric,
        "text_cells_changed": len(text_changes),
        "expected_text_cells_changed": expected_text_changes,
        "cross_representation_mentions": cross_representation_mentions,
        "independent_validation_mentions": independent_validation_mentions,
        "external_alignment_or_crosswalk_mentions": len(forbidden_external),
        "site_mapping_audit": change_counts,
        "declared_text_changes": text_changes,
    }


def file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "relative_path": relative(path),
        "sha256": sha256(path),
        "bytes": path.stat().st_size,
    }


def write_integrity(
    path: Path,
    source: Path,
    source_integrity: Path,
    output: Path,
    extra_inputs: list[Path],
    audit: dict[str, Any],
) -> None:
    if path.exists():
        raise FileExistsError(path)
    record = {
        "schema": "supplemental_workbook_scientific_integrity_v2",
        "status": "PASS",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "script_version": SCRIPT_VERSION,
        "generator": file_record(SCRIPT),
        "software": {
            "python": platform.python_version(),
            "pandas": pd.__version__,
            "openpyxl": openpyxl.__version__,
        },
        "data_integrity_policy": {
            "real_data_only": True,
            "synthetic_scientific_values": False,
            "randomness_used": False,
            "existing_numeric_cells_unchanged": True,
        },
        "source_workbook": file_record(source),
        "source_integrity": file_record(source_integrity),
        "inputs": {item.name: file_record(item) for item in extra_inputs},
        "audit": audit,
        "output": file_record(output),
    }
    path.write_text(json.dumps(record, indent=2, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")


def main() -> int:
    authenticate_inputs()
    s2_counts = build_s2()
    s3_changes = build_s3()
    s2_audit = audit_s2(s2_counts)
    s3_audit = audit_s3(s3_changes)
    write_integrity(
        S2_INTEGRITY,
        S2_SOURCE,
        S2_SOURCE_INTEGRITY,
        S2_OUTPUT,
        [ANNOTATION_CSV, ANNOTATION_MANIFEST, ANNOTATION_GENERATOR],
        s2_audit,
    )
    write_integrity(
        S3_INTEGRITY,
        S3_SOURCE,
        S3_SOURCE_INTEGRITY,
        S3_OUTPUT,
        [AEF_ALIGNMENT_MANIFEST],
        s3_audit,
    )
    print(
        json.dumps(
            {
                "table_s2_v4": {
                    "workbook": str(S2_OUTPUT.resolve()),
                    "sha256": sha256(S2_OUTPUT),
                    "integrity": str(S2_INTEGRITY.resolve()),
                    "counts": s2_counts,
                },
                "table_s3_v7": {
                    "workbook": str(S3_OUTPUT.resolve()),
                    "sha256": sha256(S3_OUTPUT),
                    "integrity": str(S3_INTEGRITY.resolve()),
                    "changes": s3_changes,
                },
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
