#!/usr/bin/env python3
"""Build an audited, versioned Table S1 without altering the original workbook.

Inputs are retained project data only. No values are simulated or imputed. The
known-faulty legacy S1E Pfam matrix is omitted from the revised workbook and is
replaced by (i) an authoritative 131-record catalog, (ii) the 126-genome
macroalgal analysis set, and (iii) a file index pointing to the regenerated raw
Pfam matrix used by the revision analyses.

Generated: 2026-07-12
Runtime: Python 3 with openpyxl 3.1.5
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
SOURCE_WORKBOOK = ROOT / "TABLES" / "Table_S1_25DEC_2025-main.xlsx"
MASTER_MANIFEST = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "integrity"
    / "reconciled_analysis_manifest_20260711_110650.csv"
)
AEF_COHORT = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "spatial"
    / "coordinate_confidence_audit_20260711_105047.csv"
)
AEF_EMBEDDINGS = (
    ROOT
    / "AlphaEarth"
    / "CSV"
    / "alphaearth_embeddings_20251019_122918.csv"
)
RAW_PFAM_MATRIX = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "analysis_stats"
    / "reconstructed_raw_pfam_counts_20260711_131706.csv.gz"
)
OUTPUT = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "supplement"
    / "Table_S1_revised_analysis_sets_20260712_180210.xlsx"
)
MANIFEST = (
    ROOT
    / "ISCIENCE_REVISION_20260711"
    / "supplement"
    / "Table_S1_revised_analysis_sets_integrity_20260712_180210.json"
)

EXPECTED_INPUT_HASHES = {
    SOURCE_WORKBOOK: None,
    MASTER_MANIFEST: None,
    AEF_COHORT: "36ab856da339da891b279bf7fffc9cbc8ccb91de37cb273d2beb9c8ba8f73da1",
    AEF_EMBEDDINGS: None,
    RAW_PFAM_MATRIX: None,
}

HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="D9EAD3")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise RuntimeError(f"Missing CSV header: {path}")
        rows = list(reader)
        return list(reader.fieldnames), rows


def style_table(ws, *, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in ws.iter_cols():
        letter = get_column_letter(column_cells[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in column_cells[: min(len(column_cells), 200)] if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def append_dict_rows(ws, columns: list[str], rows: list[dict[str, str]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])


def main() -> None:
    for path, expected in EXPECTED_INPUT_HASHES.items():
        if not path.is_file():
            raise FileNotFoundError(path)
        actual = sha256(path)
        if expected is not None and actual != expected:
            raise RuntimeError(f"Input hash mismatch for {path}: {actual} != {expected}")

    source_hash_before = sha256(SOURCE_WORKBOOK)
    master_fields, master_rows = read_csv(MASTER_MANIFEST)
    cohort_fields, cohort_rows = read_csv(AEF_COHORT)
    _, embedding_rows = read_csv(AEF_EMBEDDINGS)

    if len(master_rows) != 131 or len({row["Genome"] for row in master_rows}) != 131:
        raise RuntimeError("Authoritative master manifest is not 131 unique genomes")
    if len(cohort_rows) != 126 or len({row["genome_id"] for row in cohort_rows}) != 126:
        raise RuntimeError("The macroalgal analysis set is not 126 unique genomes")

    embedding_id_field = next(
        (field for field in embedding_rows[0] if field.lower() in {"genome", "genome_id", "genome id"}),
        None,
    )
    if embedding_id_field is None:
        raise RuntimeError("Cannot identify genome ID field in AEF embeddings")
    cohort_ids = {row["genome_id"] for row in cohort_rows}
    embedding_ids = {row[embedding_id_field] for row in embedding_rows}
    if cohort_ids != embedding_ids:
        raise RuntimeError("Analysis-set IDs differ from the retained AEF embeddings")
    if not all(row.get("clean_metadata_match", "").lower() == "true" for row in cohort_rows):
        raise RuntimeError("At least one analysis-set row does not match the retained metadata")
    if not all(row.get("aef_64_dimensions_complete", "").lower() == "true" for row in cohort_rows):
        raise RuntimeError("At least one analysis-set row lacks a complete 64-axis AEF vector")

    workbook = load_workbook(SOURCE_WORKBOOK)
    legacy_sheet = "TABLE_S1E-pfam_counts_with_met"
    if legacy_sheet not in workbook.sheetnames:
        raise RuntimeError(f"Expected legacy sheet absent: {legacy_sheet}")
    legacy_ws = workbook[legacy_sheet]
    legacy_headers = [cell.value for cell in legacy_ws[1]]
    legacy_genome_column = legacy_headers.index("Genome")
    legacy_pfam_columns = [
        index
        for index, value in enumerate(legacy_headers)
        if re.fullmatch(r"PF\d{5}", str(value))
    ]
    if not legacy_pfam_columns:
        raise RuntimeError("Legacy S1E contains no strict Pfam accession columns")
    raw_totals = {
        row["Genome"]: int(float(row["raw_pfam_hit_total"])) for row in master_rows
    }
    legacy_false_zero_genomes = []
    for values in legacy_ws.iter_rows(min_row=2, values_only=True):
        genome = str(values[legacy_genome_column]).strip()
        all_zero = all(
            float(values[index] or 0) == 0 for index in legacy_pfam_columns
        )
        if all_zero and raw_totals.get(genome, 0) > 0:
            legacy_false_zero_genomes.append(genome)
    workbook.remove(workbook[legacy_sheet])

    readme = workbook.create_sheet("README_AUDIT", 0)
    notes = [
        ("Table S1 revised audit", "Generated from retained project sources; no values were imputed or simulated."),
        ("Original workbook", str(SOURCE_WORKBOOK.relative_to(ROOT))),
        ("Original SHA-256", source_hash_before),
        ("Why legacy S1E is omitted", f"The original {len(master_rows)}-row S1E contained {len(legacy_false_zero_genomes)} false all-zero Pfam vectors caused by parser/merge failures. The original workbook remains unchanged and archived; that sheet is not carried into this analysis-ready revision."),
        ("Table S1E", "Authoritative 131-record master catalog from the reconciled analysis manifest. Exons is retained as a legacy field and is not relabeled as predicted protein count."),
        ("Table S1F", "Macroalgal genome analysis set comprising 126 records with matched Pfam profiles and 64-dimensional AEF embeddings. Quantitative coordinate uncertainty is reported as unavailable where not retained."),
        ("Table S1G", "Index of the regenerated raw Pfam matrix used for revision analyses and its SHA-256 hash. Raw counts remain primary observations; ratio and BUSCO specifications are sensitivities, not corrected counts."),
        ("Analysis-set coordinate source", str(AEF_COHORT.relative_to(ROOT))),
        ("Master source", str(MASTER_MANIFEST.relative_to(ROOT))),
    ]
    readme.append(["Item", "Audit statement"])
    for item, statement in notes:
        readme.append([item, statement])
    style_table(readme)
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 120
    for row in readme.iter_rows(min_row=2):
        row[0].fill = NOTE_FILL
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    master_columns = [
        "master_row",
        "Genome",
        "Species",
        "Phylum",
        "UsedAsReference",
        "Nucleotides",
        "Exons",
        "BUSCOs-%present",
        "Climatic zone",
        "Temperature (°C)",
        "Environment",
        "Habitat",
        "DD latitude",
        "DD longitude",
        "raw_pfam_hit_total",
        "raw_distinct_pfam_count",
        "aef_present",
        "aef_id_number",
        "aef_coordinate_status",
        "safe_for_raw_pfam_analysis",
        "safe_for_aef_pfam_analysis",
        "pfam_source_path",
        "pfam_source_sha256",
    ]
    missing_master = sorted(set(master_columns) - set(master_fields))
    if missing_master:
        raise RuntimeError(f"Missing master columns: {missing_master}")
    master_ws = workbook.create_sheet("Table S1E Master Catalog")
    append_dict_rows(master_ws, master_columns, master_rows)
    style_table(master_ws)

    cohort_columns = [
        "genome_id",
        "species",
        "phylum",
        "latitude",
        "longitude",
        "habitat",
        "climatic_zone",
        "environment",
        "metadata_temperature_c",
        "coordinate_confidence_class",
        "coordinate_evidence_class",
        "coordinate_method_documentation",
        "coordinate_source_file",
        "clean_metadata_match",
        "local_geographic_location",
        "local_reference_record",
        "reference_and_locality_traceable_locally",
        "legacy_coordinate_exact_match",
        "genomes_at_identical_coordinate",
        "coordinate_pair_is_unique",
        "nearest_sample_distance_km",
        "aef_64_dimensions_complete",
        "notes",
    ]
    missing_cohort = sorted(set(cohort_columns) - set(cohort_fields))
    if missing_cohort:
        raise RuntimeError(f"Missing cohort columns: {missing_cohort}")
    cohort_for_output = []
    for row in cohort_rows:
        output_row = dict(row)
        output_row["analysis_inclusion"] = "included_126_genome_analysis_set"
        cohort_for_output.append(output_row)
    cohort_ws = workbook.create_sheet("Table S1F Analysis Set 126")
    append_dict_rows(cohort_ws, ["analysis_inclusion", *cohort_columns], cohort_for_output)
    style_table(cohort_ws)

    file_ws = workbook.create_sheet("Table S1G Raw Pfam Index")
    file_ws.append(["File role", "Relative path", "SHA-256", "Rows", "Statement"])
    file_ws.append(
        [
            "126-genome raw Pfam count matrix",
            str(RAW_PFAM_MATRIX.relative_to(ROOT)),
            sha256(RAW_PFAM_MATRIX),
            "126 genomes + header",
            "Counts are regenerated from retained HMM output. They are primary observations; no BUSCO or denominator rescaling is applied.",
        ]
    )
    style_table(file_ws)
    file_ws.column_dimensions["E"].width = 90

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)

    # Reopen and recheck the materialized workbook rather than trusting in-memory state.
    check = load_workbook(OUTPUT, read_only=True, data_only=False)
    expected_sheets = {
        "README_AUDIT",
        "Table S1E Master Catalog",
        "Table S1F Analysis Set 126",
        "Table S1G Raw Pfam Index",
    }
    if not expected_sheets.issubset(set(check.sheetnames)):
        raise RuntimeError("Revised workbook is missing one or more audited sheets")
    if legacy_sheet in check.sheetnames:
        raise RuntimeError("Known-faulty legacy S1E was unexpectedly retained")
    if check["Table S1E Master Catalog"].max_row != 132:
        raise RuntimeError("Materialized master catalog row count is not 131 + header")
    if check["Table S1F Analysis Set 126"].max_row != 127:
        raise RuntimeError("Materialized analysis-set row count is not 126 + header")
    check.close()

    if sha256(SOURCE_WORKBOOK) != source_hash_before:
        raise RuntimeError("Original Table S1 workbook changed during build")

    phylum_counts = Counter(row["phylum"] for row in cohort_rows)
    climate_counts = Counter(row["climatic_zone"] for row in cohort_rows)
    genera = {
        row["species"].strip().split()[0]
        for row in cohort_rows
        if row.get("species", "").strip()
    }
    coordinate_sites = {
        (float(row["latitude"]), float(row["longitude"])) for row in cohort_rows
    }
    manifest = {
        "generated_by": str(Path(__file__).relative_to(ROOT)),
        "generator_sha256": sha256(Path(__file__)),
        "python_version": sys.version,
        "openpyxl_version": __import__("openpyxl").__version__,
        "output": str(OUTPUT.relative_to(ROOT)),
        "output_sha256": sha256(OUTPUT),
        "inputs": {
            str(path.relative_to(ROOT)): sha256(path) for path in EXPECTED_INPUT_HASHES
        },
        "checks": {
            "original_workbook_unchanged": True,
            "known_faulty_legacy_s1e_omitted": True,
            "legacy_false_zero_pfam_vectors": len(legacy_false_zero_genomes),
            "master_unique_genomes": len(master_rows),
            "analysis_set_unique_genomes": len(cohort_rows),
            "analysis_set_id_equals_embedding_id_set": True,
            "analysis_set_core_metadata_all_exact": True,
            "aef_64_axes_all_complete": True,
            "aef_phylum_counts": dict(sorted(phylum_counts.items())),
            "aef_climate_counts": dict(sorted(climate_counts.items())),
            "aef_unique_genera": len(genera),
            "aef_exact_coordinate_sites": len(coordinate_sites),
        },
    }
    MANIFEST.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"output": str(OUTPUT), "manifest": str(MANIFEST), **manifest["checks"]}, indent=2))


if __name__ == "__main__":
    main()
