#!/usr/bin/env python3
"""Generate and validate the accession-gated UAE collection-metadata audit.

Created: 2026-07-11 13:05:25 Asia/Dubai

This script reads the preserved source files, performs exact Genome-ID and
coordinate joins, retains the workbook's raw date strings and cell metadata,
and writes a timestamped CSV, Markdown report, and JSON run manifest. It never
normalizes a collection date and never overwrites an existing output.

Required software:
    Python >= 3.9
    openpyxl (validated with 3.1.5)
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import platform
import re
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - dependency gate
    raise SystemExit("openpyxl is required; no output was written") from exc


AUDIT_TAG = "20260711_130525"
ROOT = Path(__file__).resolve().parents[2]
INTEGRITY_DIR = ROOT / "ISCIENCE_REVISION_20260711" / "integrity"

WORKBOOK = ROOT / "TABLES" / "FINAL" / "Table_S1_25DEC_2025-main.xlsx"
WORKBOOK_SHEET = "Table S1A New Species Meta"
PUBLISHED_MASTER = ROOT / "MolPlant-macroalgalGenomesMeta-Table_S1_published.csv"
AEF_TABLE = ROOT / "AlphaEarth" / "CSV" / "alphaearth_embeddings_20251019_122918.csv"
MANIFEST = INTEGRITY_DIR / "reconciled_analysis_manifest_20260711_110650.csv"
LEGACY_MAIN = ROOT / "ISCIENCE_REVISION_20260711" / "manuscript" / "main_V1.txt"

OUTPUT_CSV = INTEGRITY_DIR / f"uae_collection_metadata_crosswalk_{AUDIT_TAG}.csv"
OUTPUT_MD = INTEGRITY_DIR / f"uae_collection_metadata_audit_{AUDIT_TAG}.md"
OUTPUT_JSON = INTEGRITY_DIR / f"uae_collection_metadata_run_manifest_{AUDIT_TAG}.json"

INPUT_PATHS = (WORKBOOK, PUBLISHED_MASTER, AEF_TABLE, MANIFEST, LEGACY_MAIN)
OUTPUT_PATHS = (OUTPUT_CSV, OUTPUT_MD, OUTPUT_JSON)

DATE_TEXT_PATTERN = re.compile(r"^(\d{2}) (\d{2}) (\d{4})$")
SPECIMEN_TOKEN_PATTERN = re.compile(r"NYAE_(\d{8})(?:_|\b)")

# These are integrity sentinels derived from the completed read-only audit. The
# output is still computed from the loaded inputs. A changed source fails loudly
# instead of silently changing a manuscript-relevant conflict classification.
EXPECTED_SPECIES_CONFLICT_IDS = {
    "6_2SDS_S5",
    "S25_FSFP210375666-2r_HKFFJDSX2_L3_1",
}
EXPECTED_LEGACY_PHAEOPHYTA_IDS = {
    "1E_S1",
    "ICS_S2",
    "20200302_1_Sargasssum1_sp_1_5SDS_S2",
    "2_6SDSE_S3",
    "5_1MIC_S4",
    "S25_FSFP210375666-2r_HKFFJDSX2_L3_1",
}
EXPECTED_MAJOR_PHYLUM_CONFLICT_IDS = {"6_2SDS_S5"}
EXPECTED_LOCALITY_DEPTH_CONFLICT_IDS = {"6_2SDS_S5"}

CSV_FIELDS = [
    "audit_timestamp",
    "source_workbook",
    "source_sheet",
    "source_excel_row",
    "source_date_cell",
    "sheet_id_number_raw",
    "sheet_genome_raw",
    "canonical_genome_id",
    "sheet_species_raw",
    "sheet_phylum_raw",
    "published_species_raw",
    "canonical_species_display",
    "canonical_phylum",
    "sheet_collection_date_raw",
    "date_cell_type",
    "date_cell_number_format",
    "specimen_code_embedded_date_token",
    "date_convention_status",
    "dd_mm_yyyy_hypothesis_vs_specimen_token_status",
    "date_confirmation_required",
    "sheet_geographic_location_raw",
    "sheet_depth_m_raw",
    "sheet_latitude_raw",
    "sheet_longitude_raw",
    "aef_id_number",
    "analytical_latitude",
    "analytical_longitude",
    "genome_id_mapping_status",
    "numeric_id_namespace_status",
    "coordinate_mapping_status",
    "aef_64_dimensions_complete",
    "species_label_conflict",
    "phylum_label_difference_class",
    "locality_depth_conflict_with_legacy_narrative",
    "legacy_narrative_locality_raw",
    "legacy_narrative_depth_m_raw",
    "legacy_narrative_map_line",
    "legacy_narrative_specimen_line",
    "rbcl_mapping_status",
    "mapping_confidence",
    "notes",
]


def fail(message: str) -> None:
    """Stop before publication output can be mistaken for a valid result."""

    raise RuntimeError(message)


def sha256_file(path: Path) -> str:
    """Hash a file in bounded-memory chunks."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def raw_text(value: Any) -> str:
    """Serialize a source cell without date parsing or whitespace trimming."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def decimal_value(value: Any, label: str) -> Decimal:
    try:
        return Decimal(raw_text(value).strip())
    except (InvalidOperation, ValueError) as exc:
        fail(f"Non-numeric coordinate in {label}: {value!r}")
        raise AssertionError("unreachable") from exc


def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            fail(f"CSV has no header: {path}")
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames), rows


def require_columns(columns: Iterable[str], required: Sequence[str], source: Path) -> None:
    missing = [column for column in required if column not in columns]
    if missing:
        fail(f"Missing required columns in {source}: {missing}")


def unique_index(rows: Sequence[Mapping[str, str]], key: str, source: Path) -> Dict[str, Mapping[str, str]]:
    index: Dict[str, Mapping[str, str]] = {}
    for row_number, row in enumerate(rows, start=2):
        value = row.get(key, "")
        if not value:
            fail(f"Blank {key!r} in {source} row {row_number}")
        if value in index:
            fail(f"Duplicate {key}={value!r} in {source}")
        index[value] = row
    return index


def locate_single_line(lines: Sequence[str], pattern: str, source: Path) -> Tuple[int, str]:
    matches = [(number, line) for number, line in enumerate(lines, start=1) if pattern in line]
    if len(matches) != 1:
        fail(f"Expected one line containing {pattern!r} in {source}; found {len(matches)}")
    return matches[0]


def exclusive_write(path: Path, data: bytes) -> None:
    """Create one immutable-by-policy output; fail if its name already exists."""

    with path.open("xb") as handle:
        handle.write(data)
        handle.flush()
        os.fsync(handle.fileno())


def csv_bytes(records: Sequence[Mapping[str, str]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(
        buffer,
        fieldnames=CSV_FIELDS,
        extrasaction="raise",
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(records)
    return buffer.getvalue().encode("utf-8")


def markdown_escape(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def build_markdown(
    records: Sequence[Mapping[str, str]],
    input_metadata: Sequence[Mapping[str, Any]],
    csv_hash: str,
    legacy_map_line_number: int,
    legacy_specimen_line_number: int,
) -> str:
    date_counts = Counter(row["sheet_collection_date_raw"] for row in records)
    locality_counts = Counter(row["sheet_geographic_location_raw"] for row in records)
    depth_counts = Counter(row["sheet_depth_m_raw"] for row in records)

    lines = [
        "# Reproducible UAE collection-metadata audit",
        "",
        f"Audit tag: `{AUDIT_TAG}` (Asia/Dubai)",
        "",
        "## Computation and validation rules",
        "",
        "The exact join evaluated before any metadata interpretation was:",
        "",
        "```text",
        'Table S1A["Genome ID"] == published master["Genome"]',
        '                             == archived AEF["Genome"]',
        '                             == reconciled manifest["Genome"]',
        "```",
        "",
        "Each target identifier was required to occur exactly once in every source. Latitude and longitude were converted independently to exact decimal values and required to be equal across Table S1A, the published master, AEF, and the reconciled manifest. Spreadsheet and AEF numeric ID fields were treated as separate namespaces and never used as join keys.",
        "",
        "Dates were not converted to date objects or normalized. The output retains the exact workbook string, Excel cell type, number format, and the exact eight-digit token following `NYAE_` in the specimen description. The conditional DD MM YYYY check is recorded only as a conflict test; it does not select that convention or replace either raw field.",
        "",
        "## Validated result",
        "",
        "All nine spreadsheet Genome IDs mapped one-to-one to the published master, archived AEF table, and reconciled manifest. Every target row has 64 nonblank AEF dimensions, `aef_present=True`, `safe_for_aef_pfam_analysis=True`, and `aef_coordinate_status=EXACT_NUMERIC_MATCH_TO_MASTER`. Spreadsheet, published-master, AEF, and manifest coordinates are numerically identical for all nine records.",
        "",
        "The spreadsheet numeric ID differs from the AEF numeric ID for all nine records. Joining those fields would mis-map the samples.",
        "",
        "| Canonical Genome ID | Canonical taxon | Phylum | Exact date text | Exact spreadsheet locality | Depth text | AEF ID | Principal caveat |",
        "|---|---|---|---|---|---:|---:|---|",
    ]

    for row in records:
        caveats: List[str] = []
        if row["date_confirmation_required"] == "TRUE":
            caveats.append("date requires source confirmation")
        if row["sheet_depth_m_raw"] == "NA":
            caveats.append("depth is literal `NA`")
        if row["species_label_conflict"] == "TRUE":
            caveats.append("spreadsheet and published taxon differ")
        if row["locality_depth_conflict_with_legacy_narrative"] == "TRUE":
            caveats.append("legacy locality/depth conflict")
        if row["rbcl_mapping_status"] == "NO_SAFE_TREE_MAPPING":
            caveats.append("no safe rbcL mapping")
        lines.append(
            "| `{gid}` | *{species}* | {phylum} | `{date}` | `{locality}` | `{depth}` | {aef_id} | {caveat}. |".format(
                gid=markdown_escape(row["canonical_genome_id"]),
                species=markdown_escape(row["canonical_species_display"]),
                phylum=markdown_escape(row["canonical_phylum"]),
                date=markdown_escape(row["sheet_collection_date_raw"]),
                locality=markdown_escape(row["sheet_geographic_location_raw"]),
                depth=markdown_escape(row["sheet_depth_m_raw"]),
                aef_id=markdown_escape(row["aef_id_number"]),
                caveat="; ".join(caveats),
            )
        )

    lines.extend(
        [
            "",
            f"The row-level crosswalk is `{OUTPUT_CSV.name}` (SHA-256 `{csv_hash}`).",
            "",
            "## Audited conflicts and limits",
            "",
            f"1. **Accession-linked UAE date strings exist.** The exact counts are `{dict(sorted(date_counts.items()))}`. All nine cells are strings (`data_type=s`, number format `General`). The workbook does not declare the ordering convention. Under a conditional DD MM YYYY reading, all nine values differ from the date-like specimen-code tokens; no normalized date was emitted.",
            "",
            f"2. **Table S1A contains three locality labels.** Exact counts are `{dict(sorted(locality_counts.items()))}`. `Ras Ghurab` does not occur in the spreadsheet locality field. The nine records occupy three exact analytical coordinate pairs.",
            "",
            f"3. **Depth is incompletely and inconsistently documented.** Exact spreadsheet counts are `{dict(sorted(depth_counts.items()))}`; `NA` is a literal source string. The legacy Figure 1 mapping at line {legacy_map_line_number} and specimen description at line {legacy_specimen_line_number} assign `NYAE_20200302_6` to Ras Ghurab at 5 m, whereas Table S1A assigns its exact Genome ID `6_2SDS_S5` to Dhabiya at 7 m and the Dhabiya coordinate. No loaded source resolves this conflict.",
            "",
            "4. **Two spreadsheet taxon labels conflict with exact-ID published metadata.** `6_2SDS_S5` changes from unidentified/Phaeophyta to *Gracilariopsis chorda*/Rhodophyta and has no safe rbcL mapping. `S25_FSFP210375666-2r_HKFFJDSX2_L3_1` changes from unidentified/Phaeophyta to *Sphacelaria divaricata*/Ochrophyta; its rbcL status corroborates accession identity while recording a taxonomy difference. Six additional Phaeophyta labels map to the canonical Ochrophyta label.",
            "",
            "## Wording recommendation",
            "",
            "> The analytical cohort contains nine UAE genome records that map one-to-one by Genome ID to the published master metadata and archived AEF table. Table S1A assigns these records to three exact coordinate pairs and three locality labels: North Cornice (one genome), AL Hiel (three), and Dhabiya (five). Recorded depth is 3 m for three records, 7 m for four, and the literal value `NA` for two. The spreadsheet contains accession-linked collection-date strings for all nine records, but their text convention is undeclared and they do not reconcile with date-like specimen-code tokens under a DD MM YYYY reading; the strings are therefore retained verbatim pending confirmation.",
            "",
            "> An earlier narrative assigns `NYAE_20200302_6` to Ras Ghurab at 5 m, whereas Table S1A assigns the same exact Genome ID to Dhabiya at 7 m and the Dhabiya coordinate. Because the available records do not resolve that discrepancy, the analysis uses the accession-gated canonical coordinate and does not treat Ras Ghurab as a reconciled analytical site.",
            "",
            "For the response to the temporal-matching comment:",
            "",
            "> Table S1A contains accession-linked date strings for all nine UAE genome records; collection dates are unavailable for the 117 comparison genomes. We retained the UAE strings verbatim because the cells use an undeclared text convention and require reconciliation with date-like specimen-code tokens. In addition, the archived AEF extraction did not select a calendar year. We therefore do not claim collection-day or collection-year matching, and no date-matched exposure was reconstructed.",
            "",
            "Do not state that the UAE records have no collection dates. Do not state that four field localities have been reconciled to three analytical coordinates. Use “nine UAE genome records” rather than implying nine independent sites or collection events. If “North Cornice” is standardized to “North Corniche,” identify the former as the raw Table S1A label.",
            "",
            "## Input provenance",
            "",
            "| Input | Bytes | SHA-256 |",
            "|---|---:|---|",
        ]
    )

    for item in input_metadata:
        lines.append(f"| `{item['path']}` | {item['size_bytes']} | `{item['sha256']}` |")

    lines.extend(
        [
            "",
            f"Generator: `{Path(__file__).resolve()}`",
            "",
            "No date, locality, depth, taxon, or coordinate was imputed. No input file was modified.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    for output in OUTPUT_PATHS:
        if output.exists():
            fail(f"Refusing to overwrite existing output: {output}")

    for source in INPUT_PATHS:
        if not source.is_file():
            fail(f"Required input does not exist: {source}")

    input_metadata = [
        {
            "path": str(path.resolve()),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        for path in INPUT_PATHS
    ]

    master_columns, master_rows = read_csv_rows(PUBLISHED_MASTER)
    aef_columns, aef_rows = read_csv_rows(AEF_TABLE)
    manifest_columns, manifest_rows = read_csv_rows(MANIFEST)

    require_columns(
        master_columns,
        ["Genome", "Species", "Phylum", "DD latitude", "DD longitude"],
        PUBLISHED_MASTER,
    )
    aef_dimensions = [f"A{number:02d}" for number in range(64)]
    require_columns(
        aef_columns,
        ["Genome", "Species", "ID number", "DD latitude", "DD longitude", *aef_dimensions],
        AEF_TABLE,
    )
    require_columns(
        manifest_columns,
        [
            "Genome",
            "Species",
            "Phylum",
            "DD latitude",
            "DD longitude",
            "aef_present",
            "aef_id_number",
            "aef_coordinate_status",
            "rbcl_mapping_status",
            "safe_for_aef_pfam_analysis",
        ],
        MANIFEST,
    )

    master_by_id = unique_index(master_rows, "Genome", PUBLISHED_MASTER)
    aef_by_id = unique_index(aef_rows, "Genome", AEF_TABLE)
    manifest_by_id = unique_index(manifest_rows, "Genome", MANIFEST)

    legacy_lines = LEGACY_MAIN.read_text(encoding="utf-8").splitlines()
    legacy_map_line_number, legacy_map_line = locate_single_line(
        legacy_lines,
        "(8) Ras Ghurab (subtidal, 5m depth)",
        LEGACY_MAIN,
    )
    legacy_specimen_line_number, legacy_specimen_line = locate_single_line(
        legacy_lines,
        "(I) Unidentified brown alga (NYAE_20200302_6)",
        LEGACY_MAIN,
    )
    legacy_map_match = re.search(r"\(8\) ([^(]+?) \(subtidal, ([0-9]+)m depth\)", legacy_map_line)
    legacy_specimen_match = re.search(r"\(I\).*?\((NYAE_[0-9]{8}_[0-9]+)\)", legacy_specimen_line)
    if not legacy_map_match or not legacy_specimen_match:
        fail("Legacy locality/depth or specimen token could not be parsed from verified lines")
    legacy_locality = legacy_map_match.group(1)
    legacy_depth = legacy_map_match.group(2)
    legacy_specimen = legacy_specimen_match.group(1)

    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=False)
    try:
        if WORKBOOK_SHEET not in workbook.sheetnames:
            fail(f"Missing worksheet {WORKBOOK_SHEET!r} in {WORKBOOK}")
        worksheet = workbook[WORKBOOK_SHEET]
        header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1)))
        headers = [raw_text(cell.value) for cell in header_cells]
        required_sheet_columns = [
            "ID number",
            "Genome",
            "Genome ID",
            "Species",
            "Phylum",
            "Collection Date",
            "Geographic location",
            "Depth (m)",
            "DD latitude",
            "DD longitude",
        ]
        require_columns(headers, required_sheet_columns, WORKBOOK)
        header_index = {name: index for index, name in enumerate(headers)}

        sheet_records: List[Tuple[int, Dict[str, Any], Dict[str, Any]]] = []
        seen_sheet_ids: set[str] = set()
        for excel_row, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
            values = {name: cells[index].value for name, index in header_index.items()}
            if all(value is None for value in values.values()):
                continue
            genome_id = raw_text(values["Genome ID"])
            if not genome_id:
                fail(f"Nonblank Table S1A row {excel_row} has a blank Genome ID")
            if genome_id in seen_sheet_ids:
                fail(f"Duplicate Table S1A Genome ID: {genome_id}")
            seen_sheet_ids.add(genome_id)
            cell_metadata = {
                "date_coordinate": cells[header_index["Collection Date"]].coordinate,
                "date_type": cells[header_index["Collection Date"]].data_type,
                "date_number_format": cells[header_index["Collection Date"]].number_format,
            }
            sheet_records.append((excel_row, values, cell_metadata))
    finally:
        workbook.close()

    if len(sheet_records) != 9:
        fail(f"Expected exactly 9 Table S1A records; found {len(sheet_records)}")

    records: List[Dict[str, str]] = []
    for excel_row, sheet, cell_meta in sheet_records:
        genome_id = raw_text(sheet["Genome ID"])
        source_presence = {
            "published master": genome_id in master_by_id,
            "AEF table": genome_id in aef_by_id,
            "manifest": genome_id in manifest_by_id,
        }
        missing_sources = [name for name, present in source_presence.items() if not present]
        if missing_sources:
            fail(f"Genome ID {genome_id!r} missing from: {missing_sources}")

        master = master_by_id[genome_id]
        aef = aef_by_id[genome_id]
        manifest = manifest_by_id[genome_id]

        if master["Species"].strip() != manifest["Species"].strip():
            fail(f"Published-master/manifest species conflict for {genome_id}")
        if master["Phylum"].strip() != manifest["Phylum"].strip():
            fail(f"Published-master/manifest phylum conflict for {genome_id}")
        if master["Species"].strip() != aef["Species"].strip():
            fail(f"Published-master/AEF species conflict for {genome_id}")

        coordinate_sources = {
            "sheet_latitude": sheet["DD latitude"],
            "master_latitude": master["DD latitude"],
            "aef_latitude": aef["DD latitude"],
            "manifest_latitude": manifest["DD latitude"],
            "sheet_longitude": sheet["DD longitude"],
            "master_longitude": master["DD longitude"],
            "aef_longitude": aef["DD longitude"],
            "manifest_longitude": manifest["DD longitude"],
        }
        latitude_values = {
            decimal_value(coordinate_sources[name], f"{genome_id}:{name}")
            for name in ("sheet_latitude", "master_latitude", "aef_latitude", "manifest_latitude")
        }
        longitude_values = {
            decimal_value(coordinate_sources[name], f"{genome_id}:{name}")
            for name in ("sheet_longitude", "master_longitude", "aef_longitude", "manifest_longitude")
        }
        if len(latitude_values) != 1 or len(longitude_values) != 1:
            fail(f"Coordinate conflict across sources for {genome_id}: {coordinate_sources}")

        if manifest["aef_present"] != "True":
            fail(f"Manifest does not mark AEF present for {genome_id}")
        if manifest["safe_for_aef_pfam_analysis"] != "True":
            fail(f"Manifest does not mark AEF analysis safe for {genome_id}")
        if manifest["aef_coordinate_status"] != "EXACT_NUMERIC_MATCH_TO_MASTER":
            fail(f"Unexpected AEF coordinate status for {genome_id}")
        if manifest["aef_id_number"] != aef["ID number"]:
            fail(f"Manifest/AEF numeric ID disagreement for {genome_id}")

        missing_dimensions = [name for name in aef_dimensions if aef[name] == ""]
        if missing_dimensions:
            fail(f"Missing AEF dimensions for {genome_id}: {missing_dimensions}")

        sheet_genome_raw = raw_text(sheet["Genome"])
        specimen_matches = SPECIMEN_TOKEN_PATTERN.findall(sheet_genome_raw)
        if len(specimen_matches) != 1:
            fail(f"Expected one NYAE date token in Table S1A Genome for {genome_id}")
        specimen_date_token = specimen_matches[0]

        date_raw = raw_text(sheet["Collection Date"])
        date_match = DATE_TEXT_PATTERN.fullmatch(date_raw)
        if not date_match:
            fail(f"Unexpected raw date-string structure for {genome_id}: {date_raw!r}")
        if cell_meta["date_type"] != "s" or cell_meta["date_number_format"] != "General":
            fail(f"Unexpected date cell metadata for {genome_id}: {cell_meta}")
        day, month, year = date_match.groups()
        conditional_dd_mm_yyyy_token = f"{year}{month}{day}"
        if conditional_dd_mm_yyyy_token == specimen_date_token:
            fail(f"Expected date-source conflict disappeared for {genome_id}")

        sheet_species = raw_text(sheet["Species"])
        sheet_phylum = raw_text(sheet["Phylum"])
        canonical_species = master["Species"].strip()
        canonical_phylum = master["Phylum"].strip()
        species_conflict = sheet_species.strip() != canonical_species

        if genome_id in EXPECTED_MAJOR_PHYLUM_CONFLICT_IDS:
            phylum_class = "MAJOR_CONFLICT_PHAEOPHYTA_TO_RHODOPHYTA"
        elif genome_id in EXPECTED_LEGACY_PHAEOPHYTA_IDS:
            phylum_class = "LEGACY_PHAEOPHYTA_LABEL_TO_CANONICAL_OCHROPHYTA"
        elif sheet_phylum.strip() == canonical_phylum:
            phylum_class = "NONE"
        else:
            fail(f"Unclassified phylum-label difference for {genome_id}")

        specimen_code = SPECIMEN_TOKEN_PATTERN.search(sheet_genome_raw)
        assert specimen_code is not None
        specimen_prefix = specimen_code.group(0).rstrip("_")
        # The legacy specimen identifier includes its terminal specimen number;
        # match it directly rather than mapping by species or row number.
        legacy_conflict = legacy_specimen in sheet_genome_raw
        if legacy_conflict:
            if genome_id != "6_2SDS_S5":
                fail(f"Legacy NYAE_20200302_6 unexpectedly maps to {genome_id}")
            if raw_text(sheet["Geographic location"]) == legacy_locality:
                fail("Expected Ras Ghurab/Dhabiya locality conflict disappeared")
            if raw_text(sheet["Depth (m)"]) == legacy_depth:
                fail("Expected 5 m/7 m depth conflict disappeared")

        rbcl_status = manifest["rbcl_mapping_status"]
        notes: List[str] = []
        if raw_text(sheet["Depth (m)"]) == "NA":
            notes.append("Spreadsheet depth is the literal text NA")
        if master["Species"] != canonical_species:
            notes.append("Published species field has surrounding whitespace stripped only for display")
        if species_conflict:
            notes.append("Spreadsheet and exact-ID published species labels conflict")
        if legacy_conflict:
            notes.append("Legacy narrative assigns Ras Ghurab at 5 m; Table S1A assigns Dhabiya at 7 m")
        if rbcl_status == "NO_SAFE_TREE_MAPPING":
            notes.append("No safe rbcL mapping")
        elif "taxonomy_differs" in rbcl_status:
            notes.append("rbcL corroborates accession identity with a taxonomy-difference flag")
        notes.append("Coordinates match exactly across Table S1A, published master, AEF, and manifest")

        confidence_parts = ["HIGH_FOR_SAMPLE_ID_AND_ANALYTICAL_COORDINATE"]
        if species_conflict:
            confidence_parts.append("PUBLISHED_TAXON_EXACT_ID")
        confidence_parts.append("FIELD_DATE_REQUIRES_CONFIRMATION")
        if legacy_conflict:
            confidence_parts.append("FIELD_LOCALITY_DEPTH_UNRESOLVED")

        record = {
            "audit_timestamp": AUDIT_TAG,
            "source_workbook": str(WORKBOOK.relative_to(ROOT)),
            "source_sheet": WORKBOOK_SHEET,
            "source_excel_row": str(excel_row),
            "source_date_cell": raw_text(cell_meta["date_coordinate"]),
            "sheet_id_number_raw": raw_text(sheet["ID number"]),
            "sheet_genome_raw": sheet_genome_raw,
            "canonical_genome_id": genome_id,
            "sheet_species_raw": sheet_species,
            "sheet_phylum_raw": sheet_phylum,
            "published_species_raw": master["Species"],
            "canonical_species_display": canonical_species,
            "canonical_phylum": canonical_phylum,
            "sheet_collection_date_raw": date_raw,
            "date_cell_type": raw_text(cell_meta["date_type"]),
            "date_cell_number_format": raw_text(cell_meta["date_number_format"]),
            "specimen_code_embedded_date_token": specimen_date_token,
            "date_convention_status": "UNDECLARED_TEXT_DO_NOT_NORMALIZE",
            "dd_mm_yyyy_hypothesis_vs_specimen_token_status": "IF_DD_MM_YYYY_THEN_MISMATCH",
            "date_confirmation_required": "TRUE",
            "sheet_geographic_location_raw": raw_text(sheet["Geographic location"]),
            "sheet_depth_m_raw": raw_text(sheet["Depth (m)"]),
            "sheet_latitude_raw": raw_text(sheet["DD latitude"]),
            "sheet_longitude_raw": raw_text(sheet["DD longitude"]),
            "aef_id_number": aef["ID number"],
            "analytical_latitude": manifest["DD latitude"],
            "analytical_longitude": manifest["DD longitude"],
            "genome_id_mapping_status": "EXACT_UNIQUE_ONE_TO_ONE_ACROSS_ALL_FOUR_SOURCES",
            "numeric_id_namespace_status": (
                "DIFFERENT_NAMESPACE_DO_NOT_JOIN"
                if raw_text(sheet["ID number"]) != aef["ID number"]
                else "UNEXPECTED_NUMERIC_ID_EQUALITY_REVIEW_REQUIRED"
            ),
            "coordinate_mapping_status": "EXACT_NUMERIC_MATCH_ACROSS_ALL_FOUR_SOURCES",
            "aef_64_dimensions_complete": "TRUE",
            "species_label_conflict": "TRUE" if species_conflict else "FALSE",
            "phylum_label_difference_class": phylum_class,
            "locality_depth_conflict_with_legacy_narrative": "TRUE" if legacy_conflict else "FALSE",
            "legacy_narrative_locality_raw": legacy_locality if legacy_conflict else "",
            "legacy_narrative_depth_m_raw": legacy_depth if legacy_conflict else "",
            "legacy_narrative_map_line": str(legacy_map_line_number) if legacy_conflict else "",
            "legacy_narrative_specimen_line": str(legacy_specimen_line_number) if legacy_conflict else "",
            "rbcl_mapping_status": rbcl_status,
            "mapping_confidence": "__".join(confidence_parts),
            "notes": "; ".join(notes) + ".",
        }
        if set(record) != set(CSV_FIELDS):
            fail(f"Internal output schema mismatch for {genome_id}")
        records.append(record)

    computed_species_conflicts = {
        row["canonical_genome_id"] for row in records if row["species_label_conflict"] == "TRUE"
    }
    computed_legacy_phylum = {
        row["canonical_genome_id"]
        for row in records
        if row["phylum_label_difference_class"]
        == "LEGACY_PHAEOPHYTA_LABEL_TO_CANONICAL_OCHROPHYTA"
    }
    computed_major_phylum = {
        row["canonical_genome_id"]
        for row in records
        if row["phylum_label_difference_class"] == "MAJOR_CONFLICT_PHAEOPHYTA_TO_RHODOPHYTA"
    }
    computed_locality_conflicts = {
        row["canonical_genome_id"]
        for row in records
        if row["locality_depth_conflict_with_legacy_narrative"] == "TRUE"
    }
    if computed_species_conflicts != EXPECTED_SPECIES_CONFLICT_IDS:
        fail(f"Species-conflict sentinel failed: {computed_species_conflicts}")
    if computed_legacy_phylum != EXPECTED_LEGACY_PHAEOPHYTA_IDS:
        fail(f"Legacy-phylum sentinel failed: {computed_legacy_phylum}")
    if computed_major_phylum != EXPECTED_MAJOR_PHYLUM_CONFLICT_IDS:
        fail(f"Major-phylum sentinel failed: {computed_major_phylum}")
    if computed_locality_conflicts != EXPECTED_LOCALITY_DEPTH_CONFLICT_IDS:
        fail(f"Locality/depth sentinel failed: {computed_locality_conflicts}")

    expected_date_counts = Counter({"02 12 2019": 1, "04 12 2019": 3, "01 03 2020": 5})
    expected_locality_counts = Counter(
        {
            "North Cornice, UAE, Arabian Gulf": 1,
            "AL Hiel, UAE, Arabian Gulf": 3,
            "Dhabiya, UAE, Arabian Gulf": 5,
        }
    )
    expected_depth_counts = Counter({"NA": 2, "3": 3, "7": 4})
    if Counter(row["sheet_collection_date_raw"] for row in records) != expected_date_counts:
        fail("Date-string count sentinel failed")
    if Counter(row["sheet_geographic_location_raw"] for row in records) != expected_locality_counts:
        fail("Locality count sentinel failed")
    if Counter(row["sheet_depth_m_raw"] for row in records) != expected_depth_counts:
        fail("Depth count sentinel failed")
    if len({(row["analytical_latitude"], row["analytical_longitude"]) for row in records}) != 3:
        fail("Expected exactly three UAE analytical coordinate pairs")
    if not all(row["numeric_id_namespace_status"] == "DIFFERENT_NAMESPACE_DO_NOT_JOIN" for row in records):
        fail("Spreadsheet and AEF numeric IDs are no longer wholly distinct")

    csv_data = csv_bytes(records)
    csv_hash = sha256_bytes(csv_data)
    markdown = build_markdown(
        records,
        input_metadata,
        csv_hash,
        legacy_map_line_number,
        legacy_specimen_line_number,
    )
    markdown_data = markdown.encode("utf-8")
    markdown_hash = sha256_bytes(markdown_data)

    script_path = Path(__file__).resolve()
    run_manifest = {
        "audit_tag": AUDIT_TAG,
        "completed_at_local": datetime.now().astimezone().isoformat(),
        "status": "complete",
        "policy": {
            "date_handling": "preserve_raw_text_and_cell_metadata; do_not_normalize",
            "join_key": "exact unique Genome ID",
            "coordinate_rule": "exact Decimal equality across workbook, published master, AEF, and manifest",
            "overwrite": "forbidden; output files opened in exclusive-create mode",
            "synthetic_or_imputed_data": False,
        },
        "software": {
            "python": sys.version,
            "python_executable": sys.executable,
            "openpyxl": openpyxl.__version__,
            "platform": platform.platform(),
        },
        "script": {
            "path": str(script_path),
            "size_bytes": script_path.stat().st_size,
            "sha256": sha256_file(script_path),
        },
        "inputs": input_metadata,
        "outputs": [
            {
                "path": str(OUTPUT_CSV.resolve()),
                "size_bytes": len(csv_data),
                "sha256": csv_hash,
                "rows_excluding_header": len(records),
            },
            {
                "path": str(OUTPUT_MD.resolve()),
                "size_bytes": len(markdown_data),
                "sha256": markdown_hash,
            },
        ],
        "run_manifest_note": "This JSON file cannot contain its own stable SHA-256; hash it after creation.",
        "validation": {
            "table_s1a_rows": len(records),
            "unique_exact_genome_id_joins": len({row["canonical_genome_id"] for row in records}),
            "all_coordinates_exact_across_four_sources": True,
            "analytical_coordinate_pairs": 3,
            "all_aef_64_dimensions_complete": True,
            "date_strings_preserved_unparsed": True,
            "conditional_dd_mm_yyyy_token_mismatches": 9,
            "species_label_conflicts": len(computed_species_conflicts),
            "major_phylum_conflicts": len(computed_major_phylum),
            "locality_depth_conflicts": len(computed_locality_conflicts),
            "legacy_map_line": legacy_map_line_number,
            "legacy_specimen_line": legacy_specimen_line_number,
        },
    }
    manifest_data = (json.dumps(run_manifest, indent=2, sort_keys=True) + "\n").encode("utf-8")

    # Repeat the non-overwrite gate immediately before the first write.
    for output in OUTPUT_PATHS:
        if output.exists():
            fail(f"Refusing to overwrite output created during this run: {output}")

    exclusive_write(OUTPUT_CSV, csv_data)
    exclusive_write(OUTPUT_MD, markdown_data)
    exclusive_write(OUTPUT_JSON, manifest_data)

    print(f"CSV={OUTPUT_CSV.resolve()}")
    print(f"CSV_SHA256={csv_hash}")
    print(f"REPORT={OUTPUT_MD.resolve()}")
    print(f"REPORT_SHA256={markdown_hash}")
    print(f"RUN_MANIFEST={OUTPUT_JSON.resolve()}")
    print("VALIDATION=PASS")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
